"""
Enhanced report generator for the Kenyan Stock Analyzer.

Generates:
  - Per-stock HTML/PDF reports with charts and signals
  - Market summary HTML/PDF with sector performance and breadth
  - Excel export with multi-sheet workbook

Charts built with matplotlib + seaborn, embedded as base64 PNG.
"""

import os
import re
import sys
import base64
import io
import json
from datetime import datetime
import logging

# Fix WeasyPrint on macOS
if sys.platform == 'darwin':
    homebrew_lib = '/opt/homebrew/lib'
    if os.path.isdir(homebrew_lib):
        existing = os.environ.get('DYLD_LIBRARY_PATH', '')
        os.environ['DYLD_LIBRARY_PATH'] = (
            f"{homebrew_lib}:{existing}" if existing else homebrew_lib
        )

import matplotlib
matplotlib.use('Agg')

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import seaborn as sns
import numpy as np
import pandas as pd
from jinja2 import Environment, FileSystemLoader
from fundamental_analysis import FundamentalAnalysis

logger = logging.getLogger(__name__)

# ---- WeasyPrint (graceful) ----
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except (OSError, ImportError) as e:
    WEASYPRINT_AVAILABLE = False
    logger.warning(
        f"WeasyPrint not available — PDF generation disabled. "
        f"Install system deps: brew install pango glib. Error: {e}"
    )

# ---- openpyxl (graceful) ----
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---- Seaborn style ----
sns.set_style("whitegrid")
sns.set_palette("muted")

# Chart colors
COLORS = {
    'price': '#2563eb',
    'sma20': '#f59e0b',
    'sma50': '#ef4444',
    'bb': '#8b5cf6',
    'rsi': '#7c3aed',
    'macd': '#2563eb',
    'macd_signal': '#ef4444',
    'macd_hist': '#9ca3af',
    'volume': '#3b82f6',
    'stoch_k': '#2563eb',
    'stoch_d': '#ef4444',
    'atr': '#8b5cf6',
    'obv': '#059669',
    'bullish': '#22c55e',
    'bearish': '#ef4444',
}


class ReportGenerator:
    """Generates HTML/PDF reports and Excel exports."""

    def __init__(self, template_dir=None, output_dir=None, clean_old=True):
        if template_dir is None:
            template_dir = os.path.join(
                os.path.dirname(__file__), '..', 'templates'
            )
        if output_dir is None:
            output_dir = os.path.join(
                os.path.dirname(__file__), '..', 'reports'
            )

        self.template_dir = os.path.abspath(template_dir)
        self.output_dir = os.path.abspath(output_dir)

        os.makedirs(self.template_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)

        # Clean old reports from previous runs (keep directory, just clear
        # files). Skipped when clean_old=False so a summary-only run does not
        # wipe an existing dashboard sitting in the same folder.
        if clean_old:
            self._clean_old_reports()

        self.env = Environment(loader=FileSystemLoader(self.template_dir))
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        logger.info(
            f"ReportGenerator: templates={self.template_dir}, "
            f"output={self.output_dir}, weasyprint={WEASYPRINT_AVAILABLE}"
        )

    def _clean_old_reports(self):
        """Remove all old report files from previous runs."""
        removed = 0
        for ext in ['.html', '.pdf', '.xlsx']:
            for fname in os.listdir(self.output_dir):
                if fname.endswith(ext):
                    try:
                        os.remove(os.path.join(self.output_dir, fname))
                        removed += 1
                    except OSError:
                        pass
        if removed > 0:
            logger.info(f"Cleaned {removed} old report files")

    # ============================================================
    #  Chart builders
    # ============================================================

    def _fig_to_b64(self):
        """Save current figure to base64 PNG string."""
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        data = base64.b64encode(buf.read()).decode('utf-8')
        plt.close()
        return data

    def _make_price_chart(self, data, title):
        """Price + SMA + Bollinger Bands."""
        fig, ax = plt.subplots(figsize=(12, 6))
        ax.plot(data.index, data['close'], label='Close', color=COLORS['price'], linewidth=1.2)
        if 'sma_20' in data.columns:
            ax.plot(data.index, data['sma_20'], label='SMA 20', color=COLORS['sma20'], linestyle='--', linewidth=1)
        if 'sma_50' in data.columns:
            ax.plot(data.index, data['sma_50'], label='SMA 50', color=COLORS['sma50'], linestyle='--', linewidth=1)
        if 'bb_upper' in data.columns and 'bb_lower' in data.columns:
            ax.plot(data.index, data['bb_upper'], color=COLORS['bb'], linestyle=':', alpha=0.6, linewidth=0.8)
            ax.plot(data.index, data['bb_lower'], color=COLORS['bb'], linestyle=':', alpha=0.6, linewidth=0.8)
            ax.fill_between(data.index, data['bb_upper'], data['bb_lower'], alpha=0.08, color=COLORS['bb'])

        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(loc='upper left', frameon=True, fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_rsi_chart(self, data, title):
        """RSI with overbought/oversold zones."""
        fig, ax = plt.subplots(figsize=(12, 3.5))
        ax.plot(data.index, data['rsi'], label='RSI', color=COLORS['rsi'], linewidth=1.2)
        ax.axhline(y=70, color=COLORS['sma50'], linestyle='--', alpha=0.7, linewidth=0.8,
                   label='Overbought (70) — may pull back')
        ax.axhline(y=30, color=COLORS['bullish'], linestyle='--', alpha=0.7, linewidth=0.8,
                   label='Oversold (30) — may bounce')
        ax.fill_between(data.index, 70, 100, alpha=0.06, color=COLORS['sma50'])
        ax.fill_between(data.index, 0, 30, alpha=0.06, color=COLORS['bullish'])
        ax.set_ylim(0, 100)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_macd_chart(self, data, title):
        """MACD, signal line, histogram."""
        fig, ax = plt.subplots(figsize=(12, 4))
        ax.plot(data.index, data['macd'], label='MACD', color=COLORS['macd'], linewidth=1.2)
        ax.plot(data.index, data['macd_signal'], label='Signal', color=COLORS['macd_signal'], linewidth=1)
        colors = [COLORS['bullish'] if v >= 0 else COLORS['sma50']
                  for v in data['macd_hist'].fillna(0)]
        ax.bar(data.index, data['macd_hist'], label='Histogram', color=colors, alpha=0.6, width=0.8)
        ax.axhline(y=0, color='black', linestyle='-', alpha=0.2, linewidth=0.5)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_volume_chart(self, data, title):
        """Volume bars with SMA overlay."""
        fig, ax = plt.subplots(figsize=(12, 3.5))
        colors = [
            COLORS['bullish'] if data['close'].iloc[i] >= data['close'].iloc[i - 1]
            else COLORS['sma50']
            for i in range(1, len(data))
        ]
        colors = [COLORS['volume']] + colors  # align length
        ax.bar(data.index, data['volume'], color=colors, alpha=0.7, width=0.8)
        # Explain the bar colours (green = price rose that day, red = fell)
        from matplotlib.patches import Patch
        legend_handles = [
            Patch(facecolor=COLORS['bullish'], alpha=0.7, label='Up day (price rose)'),
            Patch(facecolor=COLORS['sma50'], alpha=0.7, label='Down day (price fell)'),
        ]
        if 'volume_sma_20' in data.columns:
            sma_line, = ax.plot(data.index, data['volume_sma_20'], label='Vol SMA 20 (average)',
                                color=COLORS['sma20'], linewidth=1.5)
            legend_handles.append(sma_line)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(handles=legend_handles, fontsize=9)
        ax.grid(True, alpha=0.3)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(
            lambda x, p: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'
        ))
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_stochastic_chart(self, data, title):
        """Stochastic Oscillator (%K and %D)."""
        fig, ax = plt.subplots(figsize=(12, 3.5))
        ax.plot(data.index, data['stoch_k'], label='%K (fast line)', color=COLORS['stoch_k'], linewidth=1.2)
        ax.plot(data.index, data['stoch_d'], label='%D (slow line)', color=COLORS['stoch_d'], linewidth=1)
        ax.axhline(y=80, color=COLORS['sma50'], linestyle='--', alpha=0.5, linewidth=0.8,
                   label='Overbought (80) — near top of range')
        ax.axhline(y=20, color=COLORS['bullish'], linestyle='--', alpha=0.5, linewidth=0.8,
                   label='Oversold (20) — near bottom of range')
        ax.fill_between(data.index, 80, 100, alpha=0.06, color=COLORS['sma50'])
        ax.fill_between(data.index, 0, 20, alpha=0.06, color=COLORS['bullish'])
        ax.set_ylim(0, 100)
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_atr_chart(self, data, title):
        """ATR (volatility) chart."""
        fig, ax = plt.subplots(figsize=(12, 3.5))
        ax.plot(data.index, data['atr'], label='ATR', color=COLORS['atr'], linewidth=1.2)
        ax.fill_between(data.index, 0, data['atr'], alpha=0.15, color=COLORS['atr'])
        ax.set_title(title, fontsize=13, fontweight='bold')
        ax.set_xlabel('')
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30)
        return self._fig_to_b64()

    def _make_sector_chart(self, sector_data):
        """Horizontal bar chart of sector performance."""
        if not sector_data:
            return None

        sectors = list(sector_data.keys())
        changes = [sector_data[s]['avg_change_pct'] for s in sectors]

        fig, ax = plt.subplots(figsize=(10, 5))
        colors = [COLORS['bullish'] if c >= 0 else COLORS['sma50'] for c in changes]
        bars = ax.barh(sectors, changes, color=colors, alpha=0.85)

        # Add value labels
        for bar, val in zip(bars, changes):
            ax.text(
                bar.get_width() + (0.1 if val >= 0 else -0.5),
                bar.get_y() + bar.get_height() / 2,
                f'{val:+.2f}%',
                va='center', fontsize=10, fontweight='bold',
                color='#166534' if val >= 0 else '#991b1b',
            )

        ax.axvline(x=0, color='black', linewidth=0.8)
        ax.set_title('Sector Performance', fontsize=13, fontweight='bold')
        ax.set_xlabel('Average Daily Change (%)')
        ax.grid(True, alpha=0.3, axis='x')
        return self._fig_to_b64()

    # ============================================================
    #  Stock Report
    # ============================================================

    def generate_stock_report(self, symbol, analysis_result, report_type='html',
                             fundamentals=None, similar_stocks=None,
                             sector_peers=None, validation=None, score=None,
                             alerts=None, sector_medians=None, usd_kes=None):
        """
        Generate a per-stock report with charts, signals, fundamentals,
        similar stocks, and plain-English explanations.

        Args:
            symbol: Stock symbol.
            analysis_result: Dict from AnalysisEngine.analyze_stock().
            report_type: 'html', 'pdf', or 'both'.
            fundamentals: Dict of fundamental metrics from FundamentalAnalysis.
            similar_stocks: List of similar stock dicts.
            sector_peers: List of sector peer dicts.

        Returns:
            str or tuple: Path(s) to generated report(s).
        """
        if not analysis_result or 'data' not in analysis_result:
            logger.error(f"Invalid analysis result for {symbol}")
            return None

        # Normalise fundamentals so any missing metric resolves to None (not a
        # Jinja "Undefined") for stocks with no fundamentals — keeps the template
        # from erroring and falling back to the plain layout.
        from collections import defaultdict
        fundamentals = defaultdict(lambda: None, fundamentals or {})

        data = analysis_result['data']
        signals = analysis_result.get('signals', {})
        latest = analysis_result.get('latest', {})
        supports = analysis_result.get('support', [])
        resistances = analysis_result.get('resistance', [])
        daily_change = analysis_result.get('daily_change_pct')

        # Default fundamentals if none provided
        if fundamentals is None:
            fundamentals = {}
        if similar_stocks is None:
            similar_stocks = []
        if sector_peers is None:
            sector_peers = []

        # Data date: prefer fundamentals date, fallback to today
        data_date = fundamentals.get('_data_date', datetime.now().strftime('%Y-%m-%d'))

        # Build charts
        charts = [
            {'title': f'{symbol} — Price & Moving Averages', 'data': self._make_price_chart(data, f'{symbol} Price Chart')},
            {'title': f'{symbol} — RSI', 'data': self._make_rsi_chart(data, f'{symbol} RSI')},
            {'title': f'{symbol} — MACD', 'data': self._make_macd_chart(data, f'{symbol} MACD')},
        ]

        if 'volume' in data.columns:
            charts.append({
                'title': f'{symbol} — Volume',
                'data': self._make_volume_chart(data, f'{symbol} Volume'),
            })

        if 'stoch_k' in data.columns and 'stoch_d' in data.columns:
            charts.append({
                'title': f'{symbol} — Stochastic',
                'data': self._make_stochastic_chart(data, f'{symbol} Stochastic'),
            })

        if 'atr' in data.columns:
            charts.append({
                'title': f'{symbol} — ATR (Volatility)',
                'data': self._make_atr_chart(data, f'{symbol} ATR'),
            })

        # Recent data table
        recent = data.tail(10).copy()
        recent_records = []
        for idx, row in recent.iterrows():
            rec = {
                'date': idx.strftime('%Y-%m-%d') if hasattr(idx, 'strftime') else str(idx),
                'open': row.get('open'),
                'high': row.get('high'),
                'low': row.get('low'),
                'close': row.get('close'),
                'volume': int(row['volume']) if pd.notna(row.get('volume')) else None,
            }
            recent_records.append(rec)

        # Recommendation text.
        # Prefer the analyst mark (1=Strong Buy .. 5=Strong Sell) when
        # available; otherwise fall back to TradingView's technical-rating
        # gauge, which covers every stock. The label states the source so
        # it is clear this is TradingView's rating, not our own advice.
        rec = fundamentals.get('recommendation')
        if rec is None:
            label, _ = FundamentalAnalysis.signal_from_tech_rating(
                fundamentals.get('tech_rating')
            )
            if label == 'N/A':
                recommendation_text = 'No analyst coverage'
            else:
                recommendation_text = f'{label} (TradingView technical rating)'
        elif rec <= 1.5:
            recommendation_text = 'Strong Buy'
        elif rec <= 2.5:
            recommendation_text = 'Buy'
        elif rec <= 3.5:
            recommendation_text = 'Hold'
        elif rec <= 4.5:
            recommendation_text = 'Sell'
        else:
            recommendation_text = 'Strong Sell'

        # Buy/Hold/Sell class for colour-coding the executive summary
        rt = recommendation_text.lower()
        if 'sell' in rt:
            rec_class = 'sell'
        elif 'buy' in rt:
            rec_class = 'buy'
        elif 'hold' in rt or 'neutral' in rt:
            rec_class = 'hold'
        else:
            rec_class = 'none'

        # Render template
        template_data = {
            'symbol': symbol,
            'generated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'data_date': data_date,
            'latest': latest,
            'signals': signals,
            'charts': charts,
            'supports': supports,
            'resistances': resistances,
            'daily_change': daily_change,
            'recent_data': recent_records,
            # Fundamentals
            'fundamentals': fundamentals,
            'similar_stocks': similar_stocks,
            'sector_peers': sector_peers,
            'recommendation_text': recommendation_text,
            'rec_class': rec_class,
            # Accuracy & enrichment
            'validation': validation or {},
            'score': score or {},
            'alerts': alerts or [],
            'sector_context': self._sector_context(fundamentals, sector_medians),
            'usd_kes': usd_kes or {},
            # Helper functions for templates
            'fmt_mcap': self._fmt_mcap,
            'fmt_currency': self._fmt_currency,
            'interpret_pe': self._interpret_pe,
            'interpret_peg': self._interpret_peg,
            'interpret_roe': self._interpret_roe,
            'interpret_de': self._interpret_de,
            'interpret_rsi': self._interpret_rsi,
            'fund_color': self._fund_color,
        }

        html_content = self._render('stock_report.html', template_data)
        return self._save_report(f"{symbol}_report", html_content, report_type)

    # ---- Formatting helpers for templates ----

    @staticmethod
    def _sector_context(fundamentals, sector_medians):
        """Valuation vs sector median for the stock report (fails safe)."""
        if not fundamentals or not sector_medians:
            return {}
        try:
            from market_context import valuation_vs_sector
            return valuation_vs_sector(fundamentals, sector_medians)
        except Exception:
            return {}

    @staticmethod
    def _fmt_mcap(value):
        """Format market cap for display."""
        if value is None or not value:
            return 'N/A'
        try:
            v = float(value)
            if v >= 1e12:
                return f"KES {v / 1e12:.2f} Trillion"
            elif v >= 1e9:
                return f"KES {v / 1e9:.2f} Billion"
            elif v >= 1e6:
                return f"KES {v / 1e6:.2f} Million"
            return f"KES {v:,.0f}"
        except (ValueError, TypeError):
            return 'N/A'

    @staticmethod
    def _fmt_currency(value):
        """Format large currency values."""
        if value is None or not value:
            return 'N/A'
        try:
            v = float(value)
            if abs(v) >= 1e12:
                return f"KES {v / 1e12:.2f}T"
            elif abs(v) >= 1e9:
                return f"KES {v / 1e9:.2f}B"
            elif abs(v) >= 1e6:
                return f"KES {v / 1e6:.2f}M"
            return f"KES {v:,.0f}"
        except (ValueError, TypeError):
            return 'N/A'

    @staticmethod
    def _interpret_pe(pe):
        if pe is None or not pe:
            return "No data available"
        try:
            pe = float(pe)
        except (ValueError, TypeError):
            return "No data available"
        if pe < 0:
            return "⚠️ The company is currently unprofitable (negative earnings)."
        if pe < 10:
            return "✅ Low valuation — the stock is priced cheaply relative to earnings. Could be a value opportunity."
        if pe < 15:
            return "✅ Fairly valued — reasonable price for the earnings generated."
        if pe < 20:
            return "📊 Moderately valued — slightly above average, typical for growing companies."
        if pe < 30:
            return "📊 Above-average valuation — investors expect good future growth."
        return "⚠️ High valuation — the market expects very strong future growth. The stock may be expensive."

    @staticmethod
    def _interpret_peg(peg):
        if peg is None or not peg:
            return "No data available"
        try:
            peg = float(peg)
        except (ValueError, TypeError):
            return "No data available"
        if peg < 0:
            return "⚠️ Negative PEG — earnings are declining."
        if peg < 0.5:
            return "✅ Very undervalued relative to growth — potentially a bargain."
        if peg < 1.0:
            return "✅ Undervalued — P/E is lower than the earnings growth rate."
        if peg < 1.5:
            return "📊 Fairly valued — P/E is roughly in line with growth."
        if peg < 2.5:
            return "📊 Slightly overvalued — P/E exceeds growth rate."
        return "⚠️ Overvalued — stock price is high relative to earnings growth."

    @staticmethod
    def _interpret_roe(roe):
        if roe is None or not roe:
            return "No data available"
        try:
            roe = float(roe)
        except (ValueError, TypeError):
            return "No data available"
        if roe < 0:
            return "⚠️ Negative ROE — the company is destroying shareholder value."
        if roe < 5:
            return "⚠️ Weak — very little profit from shareholder money."
        if roe < 10:
            return "📊 Below average — acceptable but not impressive."
        if roe < 15:
            return "📊 Average — decent returns on shareholder capital."
        if roe < 20:
            return "✅ Good — efficiently turns shareholder money into profit."
        if roe < 30:
            return "✅ Excellent — very efficient use of shareholder capital."
        return "🌟 Outstanding — extremely efficient. Verify this is sustainable."

    @staticmethod
    def _interpret_de(ratio):
        if ratio is None or not ratio:
            return "No data available"
        try:
            ratio = float(ratio)
        except (ValueError, TypeError):
            return "No data available"
        if ratio < 0:
            return "⚠️ HIGH RISK — more liabilities than assets."
        if ratio < 0.3:
            return "✅ Very conservative — uses very little debt. Low financial risk."
        if ratio < 0.7:
            return "✅ Conservative — manageable debt levels. Low to moderate risk."
        if ratio < 1.5:
            return "📊 Moderate leverage — reasonable amount of debt."
        if ratio < 3.0:
            return "⚠️ High leverage — significant debt relative to equity."
        return "🚨 Very high leverage — heavily indebted. Proceed with caution."

    @staticmethod
    def _interpret_rsi(rsi):
        if rsi is None or not rsi:
            return "No data available"
        try:
            rsi = float(rsi)
        except (ValueError, TypeError):
            return "No data available"
        if rsi > 80:
            return "🚨 Strongly overbought — price has risen very fast, high risk of pullback."
        if rsi > 70:
            return "⚠️ Overbought — may have risen too quickly, could correct."
        if rsi > 50:
            return "✅ Bullish momentum — price trending upward with moderate strength."
        if rsi > 30:
            return "📊 Bearish momentum — price trending downward."
        if rsi > 20:
            return "⚠️ Oversold — may have fallen too far, could bounce back."
        return "🚨 Strongly oversold — extreme selling, potential for sharp reversal."

    # ============================================================
    #  Market Summary
    # ============================================================

    def generate_market_summary(self, analysis_results, sector_data=None,
                                 breadth=None, report_type='html'):
        """
        Generate market summary report with sector and breadth data.

        Returns:
            str or tuple: Path(s) to generated report(s).
        """
        # Build stock summary table
        stocks = []
        gainers = []
        losers = []

        for symbol, result in analysis_results.items():
            if not result:
                continue
            latest = result.get('latest', {})
            signals = result.get('signals', {})
            chg = result.get('daily_change_pct')

            stock = {
                'symbol': symbol,
                'price': latest.get('close'),
                'rsi': latest.get('rsi'),
                'change': round(chg, 2) if chg is not None else None,
                'signal_ma': signals.get('ma_crossover', 'undefined'),
                'signal_rsi': signals.get('rsi', 'undefined'),
                'signal_macd': signals.get('macd', 'undefined'),
                'signal_bb': signals.get('bollinger', 'undefined'),
                'trend': signals.get('trend', 'undefined'),
                'overall': signals.get('overall', 'neutral'),
            }
            stocks.append(stock)

            if chg is not None:
                entry = {'symbol': symbol, 'price': latest.get('close'), 'change': round(chg, 2)}
                if chg > 0:
                    gainers.append(entry)
                elif chg < 0:
                    losers.append(entry)

        # Sort gainers/losers
        gainers.sort(key=lambda x: x['change'], reverse=True)
        losers.sort(key=lambda x: x['change'])

        # Sector chart
        sector_chart = None
        if sector_data:
            sector_chart = self._make_sector_chart(sector_data)

        # Render
        template_data = {
            'generated_at': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'stocks': stocks,
            'gainers': gainers,
            'losers': losers,
            'sectors': sector_data,
            'breadth': breadth,
            'sector_chart': sector_chart,
        }

        html_content = self._render('market_summary.html', template_data)
        return self._save_report('market_summary', html_content, report_type)

    # ============================================================
    #  Excel Export
    # ============================================================

    def export_to_excel(self, analysis_results, sector_data=None, breadth=None):
        """
        Export analysis data to a multi-sheet Excel workbook.

        Returns:
            str: Path to the generated .xlsx file.
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not installed, skipping Excel export")
            return None

        wb = Workbook()

        # ---- Sheet 1: Summary ----
        ws = wb.active
        ws.title = "Summary"
        self._write_excel_header(ws, ['Symbol', 'Price', 'Change %', 'RSI', 'Trend',
                                       'MA Signal', 'MACD', 'Overall'])
        for i, (symbol, result) in enumerate(analysis_results.items(), 2):
            if not result:
                continue
            latest = result.get('latest', {})
            signals = result.get('signals', {})
            ws.cell(row=i, column=1, value=symbol)
            ws.cell(row=i, column=2, value=latest.get('close'))
            ws.cell(row=i, column=3, value=result.get('daily_change_pct'))
            ws.cell(row=i, column=4, value=latest.get('rsi'))
            ws.cell(row=i, column=5, value=signals.get('trend', ''))
            ws.cell(row=i, column=6, value=signals.get('ma_crossover', ''))
            ws.cell(row=i, column=7, value=signals.get('macd', ''))
            ws.cell(row=i, column=8, value=signals.get('overall', ''))

        self._auto_width(ws)

        # ---- Sheet 2: Signals ----
        ws2 = wb.create_sheet("Signals")
        signal_names = ['ma_crossover', 'rsi', 'macd', 'bollinger', 'trend',
                         'stochastic', 'volume', 'overall']
        self._write_excel_header(ws2, ['Symbol'] + [s.replace('_', ' ').title() for s in signal_names])
        for i, (symbol, result) in enumerate(analysis_results.items(), 2):
            if not result:
                continue
            signals = result.get('signals', {})
            ws2.cell(row=i, column=1, value=symbol)
            for j, sname in enumerate(signal_names, 2):
                ws2.cell(row=i, column=j, value=signals.get(sname, ''))

        self._auto_width(ws2)

        # ---- Sheet 3: Sector Analysis ----
        if sector_data:
            ws3 = wb.create_sheet("Sector Analysis")
            self._write_excel_header(ws3, ['Sector', 'Stocks', 'Avg Change %',
                                            'Avg RSI', 'Bullish %', 'Symbols'])
            for i, (sector, data) in enumerate(sector_data.items(), 2):
                ws3.cell(row=i, column=1, value=sector)
                ws3.cell(row=i, column=2, value=data['count'])
                ws3.cell(row=i, column=3, value=data['avg_change_pct'])
                ws3.cell(row=i, column=4, value=data['avg_rsi'])
                ws3.cell(row=i, column=5, value=data['bullish_ratio'])
                ws3.cell(row=i, column=6, value=', '.join(data['symbols']))
            self._auto_width(ws3)

        # Save
        path = os.path.join(
            self.output_dir, f"nse_analysis_{self.timestamp}.xlsx"
        )
        wb.save(path)
        logger.info(f"Excel exported to {path}")
        return path

    def _write_excel_header(self, ws, headers):
        """Write formatted header row."""
        header_fill = PatternFill(start_color='1e293b', end_color='1e293b', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

    def _auto_width(self, ws):
        """Auto-fit column widths."""
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    # ============================================================
    #  Helpers
    # ============================================================

    # ============================================================
    #  Summary report (compact 1-page PDF for daily email)
    # ============================================================

    def generate_summary(self, analysis_results, fundamentals_data=None,
                         validations=None, scores=None, alerts=None,
                         breadth=None, sector_data=None, usd_kes=None,
                         watchlist=None, report_type='pdf'):
        """
        Build a compact, one-page summary of key metrics and render it to PDF
        (and/or HTML). Designed for a daily email — a subset of the dashboard.

        Returns the path(s) from _save_report (str or tuple).
        """
        fundamentals_data = fundamentals_data or {}
        validations = validations or {}
        scores = scores or {}
        alerts = alerts or {}
        now = datetime.now().strftime('%Y-%m-%d %H:%M EAT')

        total = len([r for r in analysis_results.values() if r])
        bullish = sum(1 for r in analysis_results.values()
                      if r and r.get('signals', {}).get('overall') == 'bullish')
        bearish = sum(1 for r in analysis_results.values()
                      if r and r.get('signals', {}).get('overall') == 'bearish')

        # Data quality
        v_ok = sum(1 for v in validations.values() if v.get('status') == 'ok')
        v_mis = sum(1 for v in validations.values() if v.get('status') == 'mismatch')
        v_stale = sum(1 for v in validations.values() if v.get('status') == 'stale')

        fx_str = f"USD/KES {usd_kes['rate']:.2f}" if usd_kes and usd_kes.get('rate') else ''

        # ---- Watchlist rows ----
        watchlist = watchlist or sorted(analysis_results.keys())
        wl_rows = ''
        for sym in watchlist:
            r = analysis_results.get(sym)
            if not r:
                continue
            latest = r.get('latest', {})
            fund = fundamentals_data.get(sym, {})
            label, cls = FundamentalAnalysis.signal_from_tech_rating(fund.get('tech_rating'))
            chg = r.get('daily_change_pct')
            chg_str = f"{chg:+.2f}%" if chg is not None else '—'
            chg_col = '#16a34a' if (chg or 0) >= 0 else '#dc2626'
            price = latest.get('close')
            price_str = f"{price:.2f}" if price else '—'
            dy = fund.get('dividend_yield')
            dy_str = f"{dy:.1f}%" if dy else '—'
            score = scores.get(sym, {}).get('overall')
            score_str = str(score) if score is not None else '—'
            sig_col = {'strong_buy': '#16a34a', 'buy': '#16a34a', 'neutral': '#d97706',
                       'sell': '#dc2626', 'strong_sell': '#dc2626'}.get(cls, '#94a3b8')
            wl_rows += (
                f'<tr><td><b>{sym}</b></td>'
                f'<td style="color:{sig_col};font-weight:700;">{label}</td>'
                f'<td>{price_str}</td>'
                f'<td style="color:{chg_col};">{chg_str}</td>'
                f'<td>{dy_str}</td>'
                f'<td>{score_str}</td></tr>'
            )

        # ---- Top movers ----
        changes = [(s, r['daily_change_pct']) for s, r in analysis_results.items()
                   if r and r.get('daily_change_pct') is not None]
        changes.sort(key=lambda x: x[1], reverse=True)
        gainers = changes[:5]
        losers = changes[-5:][::-1] if len(changes) >= 5 else []
        movers_rows = ''
        for label, rows in [('Gainers', gainers), ('Losers', losers)]:
            for sym, chg in rows:
                col = '#16a34a' if chg >= 0 else '#dc2626'
                movers_rows += (f'<tr><td>{label}</td><td><b>{sym}</b></td>'
                                f'<td style="color:{col};">{chg:+.2f}%</td></tr>')

        # ---- Upcoming ex-dividends (next 30 days) ----
        today = datetime.now().date()
        upcoming = []
        for sym, f in fundamentals_data.items():
            ex = f.get('dividend_ex_date')
            if not ex:
                continue
            try:
                d = datetime.strptime(ex, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue
            delta = (d - today).days
            if 0 <= delta <= 30:
                upcoming.append((delta, sym, ex, f.get('dps_fy'), f.get('dividend_yield')))
        upcoming.sort()
        exdiv_rows = ''
        for delta, sym, ex, dps, dy in upcoming:
            dps_str = f"{dps:g}" if dps else '0'
            dy_str = f"{dy:.1f}%" if dy else '—'
            when = 'today' if delta == 0 else f'in {delta}d'
            exdiv_rows += (
                f'<tr><td><b>{sym}</b></td>'
                f'<td>{dps_str}</td>'
                f'<td>{dy_str}</td>'
                f'<td style="color:#166534;font-weight:700;">{ex} ({when})</td></tr>'
            )
        if not exdiv_rows:
            exdiv_rows = '<tr><td colspan="4" style="color:#94a3b8;">None in the next 30 days</td></tr>'

        # ---- Key alerts (cap to keep it one page) ----
        alert_items = ''
        shown = 0
        for sym in sorted(alerts.keys()):
            for a in alerts[sym]:
                if any(k in a for k in ['Strong Buy', 'Strong Sell', 'Oversold',
                                        'High dividend', 'Ex-dividend', '52-week']):
                    alert_items += f'<li><b>{sym}</b>: {a}</li>'
                    shown += 1
            if shown >= 14:
                break

        html = self._build_summary_html(
            now=now, fx_str=fx_str, total=total, bullish=bullish, bearish=bearish,
            v_ok=v_ok, v_mis=v_mis, v_stale=v_stale, breadth=breadth,
            wl_rows=wl_rows, movers_rows=movers_rows, exdiv_rows=exdiv_rows,
            alert_items=alert_items,
        )
        return self._save_report('nse_summary', html, report_type)

    @staticmethod
    def _build_summary_html(now, fx_str, total, bullish, bearish, v_ok, v_mis,
                            v_stale, breadth, wl_rows, movers_rows, exdiv_rows,
                            alert_items):
        breadth = breadth or {}
        breadth_str = ' · '.join(
            f"{lbl} {breadth[k]}%" for k, lbl in
            [('pct_above_sma50', 'Above SMA50'), ('pct_bullish_macd', 'Bullish MACD'),
             ('pct_rsi_above_50', 'RSI>50')] if k in breadth
        )
        alerts_block = (f'<h2>🔔 Key Alerts</h2><ul>{alert_items}</ul>'
                        if alert_items else '')
        return f'''<!DOCTYPE html><html><head><meta charset="utf-8"><style>
@page {{ size: A4; margin: 14mm; }}
body {{ font-family: -apple-system, Segoe UI, Roboto, sans-serif; color: #1e293b; font-size: 11px; }}
h1 {{ font-size: 18px; margin: 0; }}
h2 {{ font-size: 12px; border-bottom: 2px solid #3b82f6; padding-bottom: 3px; margin: 14px 0 6px; }}
.sub {{ color: #64748b; font-size: 10px; margin-bottom: 10px; }}
.pills {{ margin: 8px 0; }}
.pill {{ display: inline-block; background: #f1f5f9; border-radius: 8px; padding: 6px 12px; margin-right: 6px; }}
.pill b {{ font-size: 15px; }}
.g {{ color: #16a34a; }} .r {{ color: #dc2626; }} .a {{ color: #d97706; }}
table {{ width: 100%; border-collapse: collapse; font-size: 10px; }}
th, td {{ padding: 4px 6px; text-align: left; border-bottom: 1px solid #e2e8f0; }}
th {{ background: #f8fafc; color: #64748b; text-transform: uppercase; font-size: 8.5px; }}
ul {{ margin: 4px 0; padding-left: 18px; }} li {{ margin: 2px 0; }}
.note {{ color: #94a3b8; font-size: 9px; margin-top: 14px; border-top: 1px solid #e2e8f0; padding-top: 6px; }}
</style></head><body>
<h1>🇰🇪 NSE Daily Summary</h1>
<div class="sub">{now}{(' · ' + fx_str) if fx_str else ''}</div>
<div class="pills">
  <span class="pill"><b>{total}</b> stocks</span>
  <span class="pill g"><b>{bullish}</b> bullish</span>
  <span class="pill r"><b>{bearish}</b> bearish</span>
  <span class="pill g"><b>{v_ok}</b> price-verified</span>
  <span class="pill r"><b>{v_mis}</b> mismatch</span>
  <span class="pill a"><b>{v_stale}</b> stale</span>
</div>
{f'<div class="sub">Breadth: {breadth_str}</div>' if breadth_str else ''}
<h2>⭐ Watchlist</h2>
<table><thead><tr><th>Symbol</th><th>TV Signal</th><th>Price</th><th>Change</th><th>Yield</th><th>Score</th></tr></thead>
<tbody>{wl_rows}</tbody></table>
<h2>📈 Top Movers</h2>
<table><thead><tr><th>Dir</th><th>Symbol</th><th>Change</th></tr></thead><tbody>{movers_rows}</tbody></table>
<h2>💵 Upcoming Ex-Dividends (next 30 days)</h2>
<table><thead><tr><th>Symbol</th><th>Div KES</th><th>Yield</th><th>Ex-Date</th></tr></thead><tbody>{exdiv_rows}</tbody></table>
{alerts_block}
<div class="note">Data: TradingView (~15 min delayed), cross-checked against afx.kwayisi.org. Prices verified where possible; treat mismatches with caution. This is a mechanical summary, not investment advice.</div>
</body></html>'''

    # ============================================================
    #  Index Dashboard (single entry point)
    # ============================================================

    def generate_index(self, analysis_results, sector_data=None, breadth=None,
                       report_files=None, fundamentals_data=None,
                       validations=None, scores=None, alerts=None, usd_kes=None):
        """
        Generate the main index.html dashboard — the single entry point.

        Links to individual stock reports and shows the full market picture
        including fundamental metrics (P/E, Market Cap) from TradingView.

        Args:
            fundamentals_data: dict from FundamentalAnalysis.fetch_all_fundamentals().

        Returns:
            str: Path to index.html.
        """
        # Build stock table rows
        stocks = []
        gainers = []
        losers = []
        for symbol, result in sorted(analysis_results.items()):
            if not result:
                continue
            latest = result.get('latest', {})
            signals = result.get('signals', {})
            chg = result.get('daily_change_pct')

            # Get fundamental data for this stock
            fund = (fundamentals_data or {}).get(symbol, {})

            # TradingView Buy/Sell signal from the technical-rating gauge
            rating_label, rating_class = FundamentalAnalysis.signal_from_tech_rating(
                fund.get('tech_rating')
            )

            stock = {
                'symbol': symbol,
                'price': latest.get('close'),
                'rsi': latest.get('rsi'),
                'change': round(chg, 2) if chg is not None else None,
                'trend': signals.get('trend', 'undefined'),
                'overall': signals.get('overall', 'neutral'),
                'ma': signals.get('ma_crossover', 'undefined'),
                'macd': signals.get('macd', 'undefined'),
                'stochastic': signals.get('stochastic', 'undefined'),
                'volume_signal': signals.get('volume', 'undefined'),
                'report_file': report_files.get(symbol, '') if report_files else '',
                # Fundamental metrics
                'pe_ratio': fund.get('pe_ratio'),
                'market_cap': fund.get('market_cap'),
                'roe': fund.get('roe'),
                'peg_ratio': fund.get('peg_ratio'),
                'sector': fund.get('sector', ''),
                # Extra fundamentals (from TradingView) for the layman detail table
                'price_to_book': fund.get('price_to_book'),
                'eps': fund.get('eps_ttm'),
                'net_margin': fund.get('net_margin'),
                'debt_to_equity': fund.get('debt_to_equity'),
                'revenue_growth': fund.get('revenue_growth_yoy'),
                # TradingView Buy/Sell signal
                'signal_label': rating_label,
                'signal_class': rating_class,
                # Dividend yield (key for income-focused NSE investors)
                'dividend_yield': fund.get('dividend_yield'),
                # Dividend amount (KES/share) and ex-dividend date
                'dps': fund.get('dps_fy'),
                'ex_date': fund.get('dividend_ex_date'),
                'ex_upcoming': fund.get('dividend_ex_date_is_upcoming'),
                'dividend_status': fund.get('dividend_status'),
                'book_closure': fund.get('dividend_book_closure'),
                'payment_date': fund.get('dividend_payment_date'),
                'dividend_type': fund.get('dividend_type'),
                # Next earnings date (used by the Next Earnings page)
                'earnings_next_date': fund.get('earnings_next_date'),
                # Transparent factor score (0-100)
                'score': (scores or {}).get(symbol, {}).get('overall'),
                # Price validation (independent cross-check + freshness)
                'validation': (validations or {}).get(symbol, {}),
            }
            stocks.append(stock)

            if chg is not None:
                entry = {'symbol': symbol, 'change': round(chg, 2)}
                if chg > 0:
                    gainers.append(entry)
                elif chg < 0:
                    losers.append(entry)

        gainers.sort(key=lambda x: x['change'], reverse=True)
        losers.sort(key=lambda x: x['change'])

        # Sector chart
        sector_chart = None
        if sector_data:
            sector_chart = self._make_sector_chart(sector_data)

        # Market stats
        bullish = sum(1 for s in stocks if s['overall'] == 'bullish')
        bearish = sum(1 for s in stocks if s['overall'] == 'bearish')
        neutral = sum(1 for s in stocks if s['overall'] == 'neutral')

        # Data date
        data_date = datetime.now().strftime('%Y-%m-%d')
        if fundamentals_data:
            for _, f in fundamentals_data.items():
                if f and f.get('_data_date'):
                    data_date = f['_data_date']
                    break

        index_path = self._build_dashboard_pages(
            stocks=stocks, gainers=gainers, losers=losers, sectors=sector_data,
            breadth=breadth, sector_chart=sector_chart, bullish=bullish,
            bearish=bearish, neutral=neutral, total=len(stocks),
            data_date=data_date, alerts=alerts, usd_kes=usd_kes,
        )
        return index_path

    # Navigation across the dashboard pages (filename, label).
    _NAV = [
        ('index.html', '🏠 Overview'),
        ('technicals.html', '📈 Technicals'),
        ('fundamentals.html', '💰 Fundamentals'),
        ('dividends.html', '💵 Dividends'),
        ('earnings.html', '📅 Next Earnings'),
        ('sectors.html', '📊 Sectors'),
        ('foreign.html', '🌍 Foreign Flows'),
        ('pulse.html', '🧭 Market Pulse'),
        ('quality.html', '✅ Data Quality'),
    ]

    def _dashboard_css(self):
        """Shared stylesheet for all dashboard pages (plain string)."""
        return """<style>
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f1f5f9; color: #1e293b; }
.container { max-width: 1400px; margin: 0 auto; padding: 20px; }
.header { background: linear-gradient(135deg, #0f172a, #1e293b); color: white; padding: 26px; border-radius: 12px; margin-bottom: 16px; text-align: center; }
.header h1 { font-size: 1.8rem; margin-bottom: 5px; }
.header .date { color: #94a3b8; font-size: 0.85rem; }
/* Nav */
.nav { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 20px; }
.nav-item { padding: 10px 16px; border-radius: 8px; background: white; color: #334155; text-decoration: none; font-weight: 600; font-size: 0.9rem; box-shadow: 0 1px 3px rgba(0,0,0,0.06); border: 2px solid transparent; }
.nav-item:hover { border-color: #93c5fd; }
.nav-item.active { background: #1e293b; color: #fff; }
.page-intro { color: #64748b; font-size: 0.9rem; margin-bottom: 16px; }
/* Stats */
.stats { display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: 12px; margin-bottom: 20px; }
.stat-card { background: white; padding: 16px; border-radius: 10px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
.stat-card .stat-value { font-size: 1.8rem; font-weight: 700; }
.stat-card .stat-label { font-size: 0.75rem; color: #64748b; text-transform: uppercase; margin-top: 4px; }
.stat-card .bullish { color: #22c55e; } .stat-card .bearish { color: #ef4444; } .stat-card .neutral { color: #f59e0b; }
/* Section */
.section { background: white; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
.section h2 { font-size: 1.1rem; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #3b82f6; display: inline-block; }
/* Table */
.table-wrap { overflow-x: auto; }
table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
th, td { padding: 8px 10px; text-align: left; border-bottom: 1px solid #e2e8f0; white-space: nowrap; }
th { background: #f8fafc; color: #64748b; font-size: 0.7rem; text-transform: uppercase; font-weight: 600; position: sticky; top: 0; }
tr:hover { background: #f8fafc; }
.stock-link { color: #3b82f6; text-decoration: none; font-weight: 600; }
.stock-link:hover { text-decoration: underline; }
/* Badges */
.badge { padding: 2px 8px; border-radius: 10px; font-size: 0.7rem; font-weight: 600; text-transform: capitalize; }
.bullish, .golden_cross, .bullish_cross, .oversold, .buy { background: #dcfce7; color: #166534; }
.bearish, .death_cross, .bearish_cross, .overbought, .sell { background: #fee2e2; color: #991b1b; }
.neutral, .within_bands, .normal { background: #fef3c7; color: #92400e; }
.strong_buy { background: #16a34a; color: #fff; }
.strong_sell { background: #dc2626; color: #fff; }
.undefined { background: #f1f5f9; color: #64748b; }
.high_volume { background: #ede9fe; color: #5b21b6; }
.low_volume { background: #f1f5f9; color: #64748b; }
/* Score chips */
.score { display: inline-block; min-width: 30px; padding: 2px 8px; border-radius: 10px; font-weight: 700; font-size: 0.75rem; text-align: center; }
.score-high { background: #dcfce7; color: #166534; }
.score-mid { background: #fef3c7; color: #92400e; }
.score-low { background: #fee2e2; color: #991b1b; }
.pv-mark { font-size: 0.75rem; cursor: help; }
/* Dividend */
.div-pay { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #ccfbf1; color: #0f766e; font-weight: 700; }
.div-unverified { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #fef3c7; color: #92400e; font-weight: 700; cursor: help; }
.div-zero { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #f1f5f9; color: #94a3b8; font-weight: 600; }
.exdate-upcoming { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #16a34a; color: #fff; font-weight: 700; }
.exdate-past { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #e0e7ff; color: #3730a3; font-weight: 600; }
.exdate-none { color: #cbd5e1; }
.cal-chip { display: inline-block; padding: 2px 8px; border-radius: 10px; font-weight: 700; font-size: 0.78rem; }
.cal-near { background: #16a34a; color: #fff; }
.cal-far { background: #fde68a; color: #92400e; }
.cal-passed { background: #fecaca; color: #991b1b; }
/* Book-closure / ex-date urgency: green=future, amber=within a week, red=passed */
.bc-future { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #16a34a; color: #fff; font-weight: 700; }
.bc-soon { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #f59e0b; color: #fff; font-weight: 700; }
.bc-passed { display: inline-block; padding: 2px 8px; border-radius: 10px; background: #dc2626; color: #fff; font-weight: 700; }
.cal-legend { display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }
.cal-h3 { font-size: 0.95rem; margin-bottom: 10px; }
.cal-count { color: #94a3b8; font-weight: 400; }
/* Alerts */
.alerts-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 10px; }
.alert-card { background: #f8fafc; border: 1px solid #e2e8f0; border-left: 3px solid #3b82f6; border-radius: 8px; padding: 10px 12px; }
.alert-card .sym { font-weight: 700; color: #3b82f6; margin-bottom: 4px; }
.alert-card .items { font-size: 0.8rem; color: #475569; line-height: 1.5; }
.dq-note { font-size: 0.8rem; color: #64748b; margin-top: 8px; }
.dq-mismatch { color: #991b1b; }
.mcap-cell { font-size: 0.8rem; color: #475569; }
.positive { color: #22c55e; font-weight: 600; }
.negative { color: #ef4444; font-weight: 600; }
.grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
.grid-3 { display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; }
.sector-card { background: #f8fafc; padding: 14px; border-radius: 8px; border: 1px solid #e2e8f0; }
.sector-card h3 { font-size: 0.9rem; margin-bottom: 4px; }
.sector-change { font-size: 1.3rem; font-weight: 700; }
.sector-detail { font-size: 0.75rem; color: #64748b; margin-top: 4px; }
.chart-img { max-width: 100%; border-radius: 8px; margin-top: 12px; }
.filter-bar { margin-bottom: 16px; display: flex; gap: 8px; flex-wrap: wrap; }
.filter-bar input { padding: 8px 12px; border: 1px solid #e2e8f0; border-radius: 6px; font-size: 0.9rem; width: 220px; }
.footer { text-align: center; padding: 20px; color: #94a3b8; font-size: 0.8rem; }
/* Plain-English explainer cards */
.explain-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 14px; margin-top: 6px; }
.explain-card { background: #f8fafc; border: 1px solid #e2e8f0; border-left: 4px solid #3b82f6; border-radius: 8px; padding: 14px 16px; }
.explain-card h4 { font-size: 0.95rem; margin-bottom: 6px; color: #1e293b; }
.explain-card p { font-size: 0.84rem; color: #475569; line-height: 1.5; margin: 5px 0; }
.explain-card .eg { color: #0f766e; }
.explain-card .good { color: #166534; }
/* Fundamental cell verdicts: green = good, red = bad */
.fgood { color: #16a34a; font-weight: 700; }
.fmid { color: #d97706; font-weight: 700; }
.fbad { color: #dc2626; font-weight: 700; }
@media (max-width: 768px) { .grid-2 { grid-template-columns: 1fr; } }
</style>"""

    def _make_foreign_flow_trend_chart(self, weeks):
        """Bar chart of net foreign flow across weeks. Green = net buy, red = net sell."""
        if not weeks:
            return None
        # Oldest → newest for a chronological chart
        rows = list(reversed(weeks))
        labels = [w['week_ending'] for w in rows]
        vals = [(w.get('aggregate') or {}).get('net_foreign_flow_kes') or 0 for w in rows]
        fig, ax = plt.subplots(figsize=(11, 3.8))
        colors = [COLORS['bullish'] if v >= 0 else COLORS['sma50'] for v in vals]
        bars = ax.bar(labels, [v / 1e6 for v in vals], color=colors, alpha=0.85, width=0.6)
        ax.axhline(0, color='#334155', linewidth=0.8)
        for bar, v in zip(bars, vals):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + (2 if v >= 0 else -8),
                    f'{v / 1e6:+.0f}M', ha='center', va='bottom' if v >= 0 else 'top',
                    fontsize=8, fontweight='bold',
                    color=COLORS['bullish'] if v >= 0 else COLORS['sma50'])
        from matplotlib.patches import Patch
        ax.legend(handles=[
            Patch(facecolor=COLORS['bullish'], alpha=0.85, label='Net foreign BUY (inflow)'),
            Patch(facecolor=COLORS['sma50'], alpha=0.85, label='Net foreign SELL (outflow)'),
        ], loc='upper left', fontsize=9)
        ax.set_title('Weekly Net Foreign Flow (KES, millions)', fontsize=13, fontweight='bold')
        ax.set_ylabel('KES millions')
        ax.grid(True, alpha=0.3, axis='y')
        plt.xticks(rotation=30, ha='right')
        return self._fig_to_b64()

    def _make_foreign_participation_chart(self, weeks):
        """Line chart of foreign participation % (share of total turnover)."""
        rows = list(reversed(weeks))
        pts = [((w['week_ending'], (w.get('aggregate') or {}).get('foreign_participation_pct')))
               for w in rows]
        pts = [(d, v) for d, v in pts if v is not None]
        if not pts:
            return None
        labels = [p[0] for p in pts]
        vals = [p[1] for p in pts]
        fig, ax = plt.subplots(figsize=(11, 3.2))
        ax.plot(labels, vals, marker='o', color=COLORS['macd'], linewidth=1.6, markersize=6,
                label='Foreign share of weekly turnover')
        for x, y in zip(labels, vals):
            ax.text(x, y + 0.8, f'{y:.1f}%', ha='center', fontsize=8, color=COLORS['macd'])
        ax.set_title('Foreign Participation (% of total NSE turnover)', fontsize=13, fontweight='bold')
        ax.set_ylabel('Foreign %')
        ax.legend(loc='upper left', fontsize=9)
        ax.grid(True, alpha=0.3)
        plt.xticks(rotation=30, ha='right')
        return self._fig_to_b64()

    def _page_shell(self, page_title, active_file, subtitle, body_html, with_filter=False):
        """Wrap a page body in the shared shell (head, header, nav, footer)."""
        nav = ''.join(
            f'<a href="{fn}" class="nav-item{" active" if fn == active_file else ""}">{lbl}</a>'
            for fn, lbl in self._NAV
        )
        js = ("<script>function filterTable(){var i=document.getElementById('search');"
              "var q=i?i.value.toLowerCase():'';var rows=document.querySelectorAll('#mainTable tbody tr');"
              "rows.forEach(function(r){r.style.display=(!q||r.textContent.toLowerCase().indexOf(q)>-1)?'':'none';});}</script>"
              ) if with_filter else ""
        return (
            '<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8">'
            '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
            f'<title>{page_title}</title>' + self._dashboard_css() + '</head><body>'
            '<div class="container">'
            f'<div class="header"><h1>🇰🇪 NSE Dashboard</h1><div class="date">{subtitle}</div></div>'
            f'<div class="nav">{nav}</div>'
            f'{body_html}'
            '<div class="footer">Generated by Kenyan Stock Analyzer · '
            'Click any stock symbol for its full individual report</div>'
            '</div>' + js + '</body></html>'
        )

    @staticmethod
    def _fund_verdict(metric, value):
        """
        Judge a fundamental value as 'good', 'bad', or None (neutral), using
        the same layman thresholds shown in the plain-English guide. Used to
        colour cells green (good) / red (bad) so quality is visible at a glance.
        """
        if value is None:
            return None
        try:
            v = float(value)
        except Exception:
            # Covers non-numeric values and Jinja Undefined (stocks with no
            # fundamentals) — treat as neutral (no colour) rather than error.
            return None
        # Three tiers: 'good' (green), 'bad' (red), 'mid' (amber). A numeric
        # value that is neither clearly good nor bad is 'mid'. Non-numeric /
        # missing values returned None above and stay uncoloured (black).
        rules = {
            'pe':    lambda: 'good' if 0 < v <= 15 else 'bad' if (v <= 0 or v > 30) else 'mid',
            'peg':   lambda: 'good' if 0 < v <= 1 else 'bad' if (v < 0 or v > 2.5) else 'mid',
            'pb':    lambda: 'good' if 0 < v <= 1.5 else 'bad' if (v < 0 or v > 5) else 'mid',
            'ps':    lambda: 'good' if 0 < v <= 1.5 else 'bad' if (v < 0 or v > 6) else 'mid',
            'ev_ebitda': lambda: 'good' if 0 < v <= 10 else 'bad' if (v < 0 or v > 20) else 'mid',
            'eps':   lambda: 'good' if v > 0 else 'bad' if v < 0 else 'mid',
            'eps_growth': lambda: 'good' if v >= 10 else 'bad' if v < 0 else 'mid',
            'roe':   lambda: 'good' if v >= 15 else 'bad' if v < 5 else 'mid',
            'roic':  lambda: 'good' if v >= 12 else 'bad' if v < 5 else 'mid',
            'roa':   lambda: 'good' if v >= 8 else 'bad' if v < 2 else 'mid',
            'gm':    lambda: 'good' if v >= 40 else 'bad' if v < 15 else 'mid',
            'om':    lambda: 'good' if v >= 15 else 'bad' if v < 0 else 'mid',
            'nm':    lambda: 'good' if v >= 15 else 'bad' if v < 0 else 'mid',
            'fcfm':  lambda: 'good' if v >= 10 else 'bad' if v < 0 else 'mid',
            'de':    lambda: 'good' if 0 <= v <= 0.5 else 'bad' if (v < 0 or v > 2) else 'mid',
            'cr':    lambda: 'good' if v >= 1.5 else 'bad' if v < 1 else 'mid',
            'qr':    lambda: 'good' if v >= 1 else 'bad' if v < 0.7 else 'mid',
            'rg':    lambda: 'good' if v >= 10 else 'bad' if v < 0 else 'mid',
            'yield': lambda: 'good' if v >= 5 else 'mid',
        }
        fn = rules.get(metric)
        return fn() if fn else None

    @staticmethod
    def _fund_color(metric, value):
        """Return a colour class for templates: green / amber / red / none.
        good→'positive' (green), mid→'midv' (amber), bad→'negative' (red),
        None (N/A)→'' (black)."""
        verdict = ReportGenerator._fund_verdict(metric, value)
        return {'good': 'positive', 'mid': 'midv', 'bad': 'negative'}.get(verdict, '')

    def _fundamentals_explainer(self):
        """Plain-English guide to each fundamental metric, for non-experts."""
        cards = [
            ("P/E — Price to Earnings",
             "How many shillings you pay for each 1 shilling of yearly profit.",
             "P/E of 10 → you pay KES 10 for every KES 1 the company earns a year (about 10 years of profit to earn back the price).",
             "Under 10 = cheap (or troubled); 10–20 = fair; over 25 = expensive (fast growth expected). Always compare within the same sector."),
            ("PEG — P/E adjusted for growth",
             "A high P/E is fine if profits grow fast. PEG divides the P/E by the growth rate to check that.",
             "P/E of 20 but profits growing 20%/yr → PEG = 1.0 (fairly priced).",
             "Under 1.0 = attractively priced for its growth; around 1 = fair; over 2 = pricey."),
            ("P/B — Price to Book",
             "Price versus the company's net worth on paper (assets minus debts) per share.",
             "P/B of 1 = you pay exactly the company's book value; 2 = twice that.",
             "Under 1 can be cheap (common for NSE banks); 1–3 is typical; high = paying a premium for brand/growth."),
            ("EPS — Earnings Per Share",
             "The company's yearly profit split across each share — the profit that 'belongs' to one share.",
             "EPS of KES 5 → each share earned 5 shillings this year. It's the 'E' in P/E.",
             "Higher and rising year on year is better. Negative EPS = the company is losing money."),
            ("ROE — Return on Equity",
             "How much profit the company squeezes out of shareholders' money.",
             "ROE of 20% → for every KES 100 of shareholder money, it makes KES 20 profit a year.",
             "15–20% = good; above 20% = excellent (check it's sustainable); below 10% = weak."),
            ("Net Margin",
             "Out of every 100 shillings of sales, how many shillings become actual profit after ALL costs and taxes.",
             "Net margin of 25% → KES 25 profit from every KES 100 of sales.",
             "10%+ = solid; 20%+ = strong. It varies a lot by industry, so compare like with like."),
            ("D/E — Debt to Equity",
             "How much the company has borrowed versus what shareholders own — its debt load and risk.",
             "D/E of 1.0 → debt equals shareholder equity; 0.3 → very little debt.",
             "Under 0.5 = conservative/safe; 0.5–1.5 = moderate; over 2 = risky. Banks naturally run higher."),
            ("Rev Growth — Revenue Growth",
             "How much total sales grew compared with a year ago — is the business getting bigger?",
             "+15% → sales are 15% higher than last year; a red −5% → sales shrank.",
             "10%+ = healthy; flat is okay for a mature company; negative is a warning sign."),
            ("Yield — Dividend Yield",
             "The yearly dividend as a percentage of the share price — your income return just for holding it.",
             "Yield of 6% → KES 6 a year for every KES 100 invested.",
             "4–8% is attractive on the NSE — but confirm it's sustainable (see the Dividends page)."),
            ("Score (0–100)",
             "Our own transparent screen that blends value, quality, momentum, dividend and liquidity into one number.",
             "85 = strong across the board; 40 = weak. Higher means stronger on these factors overall.",
             "A quick way to compare stocks at a glance — it is a mechanical guide, NOT a recommendation to buy or sell."),
        ]
        items = ''
        for title, what, eg, good in cards:
            items += (f'<div class="explain-card"><h4>{title}</h4>'
                      f'<p>{what}</p>'
                      f'<p class="eg">📌 <strong>Example:</strong> {eg}</p>'
                      f'<p class="good">✅ <strong>What\'s good:</strong> {good}</p></div>')
        return ('<div class="section"><h2>📖 What these numbers mean (plain English)</h2>'
                '<p class="page-intro">No accounting needed — here is each column explained simply, '
                'with an example and what counts as a good value.</p>'
                f'<div class="explain-grid">{items}</div></div>')

    def _build_foreign_flows_body(self, sym_td):
        """
        Build the Foreign Flows page body from the manual weekly input at
        manual_input/foreign_flows.json. Returns HTML.
        Renders a friendly empty-state if no data / file missing.
        """
        try:
            from foreign_flows import load as load_foreign
            data = load_foreign()
        except Exception as e:
            logger.warning(f"Foreign flows: loader failed: {e}")
            data = {"weeks": [], "source_home": "", "source_note": ""}

        weeks = data.get("weeks") or []
        source_home = data.get("source_home") or "https://www.nse.co.ke/market-statistics/"

        if not weeks:
            return (
                '<p class="page-intro">Foreign-investor participation in the NSE — '
                'weekly figures from the official NSE Weekly Market Statistics bulletin.</p>'
                '<div class="section"><h2>🌍 Foreign Flows</h2>'
                '<div class="dq-note dq-mismatch">⚠️ No data yet.</div>'
                '<p>To populate this page, update <code>manual_input/foreign_flows.json</code> '
                'with figures from the latest NSE Weekly Market Statistics bulletin '
                f'(<a href="{source_home}" target="_blank">source</a>). '
                'See <code>manual_input/README.md</code> for the 5-minute instructions.</p>'
                '<p class="dq-note">This is the honest state: no free automated feed publishes '
                'per-stock daily foreign activity, so we enter weekly figures by hand from the '
                "authoritative NSE bulletin rather than guess.</p></div>"
                + self._foreign_flows_glossary()
            )

        latest = weeks[0]
        agg = latest.get('aggregate') or {}
        net = agg.get('net_foreign_flow_kes')
        buys = agg.get('foreign_buys_kes')
        sells = agg.get('foreign_sells_kes')
        pct = agg.get('foreign_participation_pct')

        # ---- Headline summary (colour-coded) ----
        def _fmt_m(v):
            if v is None:
                return '—'
            return (f"KES {v / 1e9:.2f}B" if abs(v) >= 1e9
                    else f"KES {v / 1e6:.1f}M")
        net_cls = ('positive' if (net or 0) > 0 else 'negative' if (net or 0) < 0 else 'neutral')
        net_word = ('NET BUY (inflow)' if (net or 0) > 0
                    else 'NET SELL (outflow)' if (net or 0) < 0 else 'FLAT')
        summary_html = (
            '<div class="stats">'
            f'<div class="stat-card"><div class="stat-value {net_cls}">{_fmt_m(net)}</div>'
            f'<div class="stat-label">Net foreign flow</div></div>'
            f'<div class="stat-card"><div class="stat-value bullish">{_fmt_m(buys)}</div>'
            f'<div class="stat-label">Foreign BUYS</div></div>'
            f'<div class="stat-card"><div class="stat-value bearish">{_fmt_m(sells)}</div>'
            f'<div class="stat-label">Foreign SELLS</div></div>'
            f'<div class="stat-card"><div class="stat-value">{pct:.1f}%' if pct is not None else
            f'<div class="stat-card"><div class="stat-value">—'
        )
        summary_html += (
            f'</div><div class="stat-label">Foreign participation</div></div>'
            f'<div class="stat-card"><div class="stat-value" style="font-size:1.1rem;">{latest["week_ending"]}</div>'
            f'<div class="stat-label">Week ending · {net_word}</div></div>'
            '</div>'
        )

        # ---- Top buys / sells side by side with proportional bars ----
        def _stock_rows(items, colour):
            if not items:
                return '<p class="dq-note">No entries reported this week.</p>'
            max_v = max((r['value_kes'] for r in items), default=1) or 1
            body = ''
            for r in items:
                pct_bar = int((r['value_kes'] / max_v) * 100)
                body += (
                    '<tr><td><strong>' + r['symbol'] + '</strong></td>'
                    f'<td>{_fmt_m(r["value_kes"])}</td>'
                    '<td><div style="background:#e2e8f0; border-radius:4px; overflow:hidden;">'
                    f'<div style="height:14px; width:{pct_bar}%; background:{colour};"></div>'
                    '</div></td></tr>'
                )
            return ('<table><thead><tr><th>Symbol</th><th>Value</th><th>Relative size</th>'
                    '</tr></thead><tbody>' + body + '</tbody></table>')

        buys_html = _stock_rows(latest.get('top_foreign_buys'), '#16a34a')
        sells_html = _stock_rows(latest.get('top_foreign_sells'), '#dc2626')
        top_html = (
            '<div class="grid-2">'
            f'<div class="section"><h2>🟢 Top Foreign BUYS — week of {latest["week_ending"]}</h2>{buys_html}</div>'
            f'<div class="section"><h2>🔴 Top Foreign SELLS — week of {latest["week_ending"]}</h2>{sells_html}</div>'
            '</div>'
        )

        # ---- Trend charts (bar for net flow, line for participation %) ----
        chart_html = ''
        if len(weeks) >= 1:
            b64_flow = self._make_foreign_flow_trend_chart(weeks)
            b64_pct = self._make_foreign_participation_chart(weeks)
            if b64_flow:
                chart_html += (f'<div class="section"><h2>📈 Weekly Net Foreign Flow</h2>'
                               f'<img src="data:image/png;base64,{b64_flow}" class="chart-img" '
                               'alt="Weekly Net Foreign Flow">'
                               '<div class="dq-note">Green bars = weeks foreigners were net buyers · '
                               'red = net sellers. Longer bar = larger flow.</div></div>')
            if b64_pct:
                chart_html += (f'<div class="section"><h2>📊 Foreign Participation (% of NSE turnover)</h2>'
                               f'<img src="data:image/png;base64,{b64_pct}" class="chart-img" '
                               'alt="Foreign Participation %">'
                               '<div class="dq-note">Share of the total NSE weekly turnover attributable '
                               'to foreign investors. Higher = foreigners are more active in the market.</div></div>')

        # ---- History table (all weeks) ----
        hist_rows = ''
        for w in weeks:
            a = w.get('aggregate') or {}
            n = a.get('net_foreign_flow_kes')
            cls = 'positive' if (n or 0) > 0 else 'negative' if (n or 0) < 0 else ''
            p = a.get('foreign_participation_pct')
            hist_rows += (f'<tr><td><strong>{w["week_ending"]}</strong></td>'
                          f'<td>{_fmt_m(a.get("foreign_buys_kes"))}</td>'
                          f'<td>{_fmt_m(a.get("foreign_sells_kes"))}</td>'
                          f'<td class="{cls}">{_fmt_m(n)}</td>'
                          f'<td>{p:.1f}%' if p is not None else '<td>—')
            hist_rows += '</td></tr>'
        history_html = (
            '<div class="section"><h2>📜 History (all weeks)</h2>'
            '<div class="table-wrap"><table><thead><tr>'
            '<th>Week ending</th><th>Foreign buys</th><th>Foreign sells</th>'
            '<th>Net flow</th><th>Participation %</th>'
            f'</tr></thead><tbody>{hist_rows}</tbody></table></div>'
            '<div class="dq-note">Newest first. Every row here was entered by hand from the '
            f'<a href="{source_home}" target="_blank">NSE Weekly Market Statistics bulletin</a>.</div>'
            '</div>'
        )

        # ---- Source & disclaimer ----
        src_html = (
            '<div class="section"><h2>ℹ️ Source &amp; how this page works</h2>'
            f'<p>Latest figures are for the week ending <strong>{latest["week_ending"]}</strong> '
            f'per <em>{latest.get("source_label","NSE Weekly Market Statistics")}</em>'
            + (f' — <a href="{latest["source_url"]}" target="_blank">bulletin</a>' if latest.get("source_url") else '')
            + '.</p>'
            '<p class="dq-note">No free automated feed publishes per-stock daily foreign activity '
            'for the NSE. Instead of guessing, we enter the weekly figures manually from the NSE '
            'bulletin — every number on this page is traceable to that source. To refresh, edit '
            '<code>manual_input/foreign_flows.json</code> once a week; see '
            '<code>manual_input/README.md</code> for the 5-minute steps.</p></div>'
        )

        return (
            '<p class="page-intro">Who is buying and selling on the NSE — foreign investors '
            'vs. local. Positive net flow = foreigners were net buyers that week.</p>'
            + summary_html + top_html + chart_html + history_html + src_html
            + self._foreign_flows_glossary()
        )

    def _foreign_flows_glossary(self):
        """Plain-English guide for Foreign Flows terms."""
        cards = [
            ("Foreign investor",
             "An investor whose registered address is outside Kenya (individual, institution, or fund).",
             "A pension fund in London buying SCOM shares on the NSE is a foreign investor."),
            ("Foreign BUYS (KES)",
             "Total value of shares that foreign investors bought that week.",
             "Foreign investors bought KES 1.2B of NSE shares → foreign buys = 1.2B."),
            ("Foreign SELLS (KES)",
             "Total value of shares that foreign investors sold that week.",
             "Foreign investors sold KES 1.35B → foreign sells = 1.35B."),
            ("Net foreign flow",
             "Buys minus Sells. Positive = net inflow (foreigners were net buyers). Negative = net outflow (they sold more than they bought).",
             "Buys 1.2B − Sells 1.35B = Net −150M → foreign investors were net sellers."),
            ("Foreign participation %",
             "The share of total weekly NSE trading value contributed by foreigners.",
             "60% participation → foreigners were on one side of 60% of the trades by value."),
            ("Why it matters",
             "Foreign flows heavily influence NSE large-caps (SCOM, EQTY, KCB, EABL). A week of heavy foreign buying often supports prices; heavy selling often pressures them.",
             "If foreigners are consistently buying SCOM, that's usually price-supportive. Consistent selling of a stock is a headwind."),
        ]
        items = ''
        for title, what, eg in cards:
            items += (f'<div class="explain-card"><h4>{title}</h4>'
                      f'<p>{what}</p>'
                      f'<p class="eg">📌 <strong>Example:</strong> {eg}</p></div>')
        return ('<div class="section"><h2>📖 What these terms mean (plain English)</h2>'
                '<p class="page-intro">No finance degree needed — each term explained simply, with '
                'a worked example.</p>'
                f'<div class="explain-grid">{items}</div></div>')

    def _build_market_pulse_body(self, analysis_results_for_pulse=None):
        """
        Build the Market Pulse page from live sources (news, CBK, oil, FX,
        African indices). Fails safe per-source — the page renders whatever
        came back and shows a note for anything missing.
        """
        try:
            from market_pulse import load_all as load_pulse
            p = load_pulse(analysis_results_for_pulse or {})
        except Exception as e:
            logger.warning(f"Market Pulse: loader failed: {e}")
            p = {"news": [], "cbk": {}, "oil": [], "african_indices": [],
                 "fx": [], "generated_at": ""}

        def _pct_span(v):
            if v is None:
                return '<span>—</span>'
            cls = 'positive' if v >= 0 else 'negative'
            return f'<span class="{cls}">{v:+.2f}%</span>'

        # ---- 1. Monetary policy (CBK) ----
        cbk = p.get('cbk') or {}
        cbr = cbk.get('cbr_pct')
        cbk_html = (
            '<div class="section"><h2>🏦 Monetary Policy — Central Bank of Kenya</h2>'
            '<div class="stats">'
            f'<div class="stat-card"><div class="stat-value">{f"{cbr:.2f}%" if cbr is not None else "—"}</div>'
            '<div class="stat-label">Central Bank Rate (CBR)</div></div>'
            '</div>'
        )
        if cbk.get('cbr_note'):
            cbk_html += f'<p class="dq-note"><strong>MPC statement:</strong> {cbk["cbr_note"]}</p>'
        if cbk.get('inflation_note'):
            cbk_html += f'<p class="dq-note"><strong>Inflation:</strong> {cbk["inflation_note"]}</p>'
        cbk_html += (f'<p class="dq-note">Source: <a href="{cbk.get("source_url", "#")}" '
                     'target="_blank">centralbank.go.ke</a>. Higher CBR usually pressures bank '
                     "loan books but boosts their bond income; falling CBR is the reverse.</p></div>")

        # ---- 2. Oil prices ----
        oil = p.get('oil') or []
        if oil:
            cards = ''
            for o in oil:
                chg = o.get('change_1w_pct')
                # For oil: rising = bad for KES importer; label colour accordingly
                cls = 'negative' if (chg or 0) > 0 else 'positive'
                cards += (f'<div class="stat-card"><div class="stat-value {cls}">'
                          f'${o["price_usd"]:.2f}</div>'
                          f'<div class="stat-label">{o["name"]} · 1w {_pct_span(chg)}</div></div>')
            oil_html = ('<div class="section"><h2>🛢️ Oil Prices (USD/barrel)</h2>'
                        f'<div class="stats">{cards}</div>'
                        '<p class="dq-note">Kenya is a net oil importer. <strong>Rising oil (red)</strong> '
                        'pressures the KES and can lift inflation; <strong>falling oil (green)</strong> is '
                        'usually favourable. Source: Yahoo Finance.</p></div>')
        else:
            oil_html = ''

        # ---- 3. FX ----
        fx = p.get('fx') or []
        if fx:
            cards = ''
            for r in fx:
                cards += (f'<div class="stat-card"><div class="stat-value">KES {r["rate_kes"]:.2f}</div>'
                          f'<div class="stat-label">1 {r["code"]} = KES · {r["name"]}</div></div>')
            fx_html = ('<div class="section"><h2>💱 KES Exchange Rates</h2>'
                       f'<div class="stats">{cards}</div>'
                       f'<p class="dq-note">Updated: {fx[0].get("updated", "")}. '
                       'A weaker KES helps exporters (tea, coffee, tourism); a stronger KES helps '
                       'importers (fuel, machinery) and the many NSE-listed dual-currency names.</p></div>')
        else:
            fx_html = ''

        # ---- 4. African markets comparison ----
        ai = p.get('african_indices') or []
        if ai:
            rows = ''
            for idx in ai:
                d1 = idx.get('change_1d_pct')
                w1 = idx.get('change_1w_pct')
                rows += (f'<tr><td><strong>{idx["name"]}</strong></td>'
                         f'<td>{idx["country"]}</td>'
                         f'<td>{f"{idx["price"]:,.2f}" if idx.get("price") else "—"}</td>'
                         f'<td>{_pct_span(d1)}</td><td>{_pct_span(w1)}</td></tr>')
            # Kenya rank note
            valid = [(idx["name"], idx.get("change_1d_pct")) for idx in ai
                     if idx.get("change_1d_pct") is not None]
            valid.sort(key=lambda x: x[1], reverse=True)
            kenya_rank_html = ''
            for i, (name, chg) in enumerate(valid):
                if '🇰🇪' in name:
                    kenya_rank_html = (f'<p class="dq-note">🇰🇪 <strong>NSE ranks '
                                       f'#{i+1} of {len(valid)}</strong> among African markets today '
                                       f'({chg:+.2f}%).</p>')
                    break
            african_html = ('<div class="section"><h2>🌍 African Markets Today</h2>'
                            '<div class="table-wrap"><table><thead><tr>'
                            '<th>Index</th><th>Country</th><th>Price</th>'
                            '<th>1-day</th><th>1-week</th></tr></thead>'
                            f'<tbody>{rows}</tbody></table></div>'
                            + kenya_rank_html +
                            '<p class="dq-note">Source: TradingView. Kenya is derived from our own NSE stock '
                            'data (equal-weight average) since Yahoo/TV do not carry a reliable NSE 20 symbol.</p></div>')
        else:
            african_html = ''

        # ---- 5. News (headlines only, NO auto-sentiment) ----
        news = p.get('news') or []
        news_html = ''
        if news:
            # Group by topic
            by_topic = {}
            for n in news:
                by_topic.setdefault(n['topic'], []).append(n)
            blocks = ''
            for topic in ['NSE / Kenyan Stocks', 'Kenyan Banking', 'Central Bank of Kenya',
                          'Kenyan Economy', 'Oil / Global']:
                items = by_topic.get(topic, [])
                if not items:
                    continue
                cards = ''
                for n in items[:6]:
                    # Format date compactly
                    date_short = n.get('published_utc', '')
                    m = re.match(r'(\w{3}, \d{1,2} \w{3} \d{4})', date_short)
                    date_display = m.group(1) if m else date_short[:16]
                    src = n.get('source', '')
                    cards += (
                        f'<div class="alert-card">'
                        f'<div class="items"><a href="{n["url"]}" target="_blank" '
                        'style="color:#3b82f6; text-decoration:none; font-weight:600;">'
                        f'{n["title"]}</a></div>'
                        f'<div style="font-size:0.72rem; color:#94a3b8; margin-top:6px;">'
                        f'{date_display} · {src}</div></div>'
                    )
                blocks += (f'<h3 class="cal-h3" style="margin-top:14px;">📰 {topic}</h3>'
                           f'<div class="alerts-grid">{cards}</div>')
            news_html = ('<div class="section"><h2>📰 Latest Headlines</h2>'
                         + blocks +
                         '<p class="dq-note">Source: Google News RSS. Headlines are shown neutral '
                         'with date &amp; publisher — we do NOT auto-classify positive/negative '
                         '(free sentiment tools on financial text are too unreliable to trust with money). '
                         'Click a headline to read the original.</p></div>')

        # ---- 6. Foreign flows chip (links to the dedicated page) ----
        foreign_chip = (
            '<div class="section"><h2>🌍 Foreign Investor Activity</h2>'
            '<p>Weekly figures from the NSE Weekly Market Statistics bulletin are on the '
            '<a href="foreign.html" style="color:#3b82f6; font-weight:600;">Foreign Flows page →</a></p></div>'
        )

        # ---- 7. How to read this page ----
        howto = self._market_pulse_glossary()

        # ---- Assemble ----
        return (
            '<p class="page-intro">Context around the NSE trading day — the macro forces '
            'that shape prices. Everything here is <strong>refreshed on every run</strong> '
            'from named sources; nothing is auto-classified into a "sentiment" (see the note '
            'at the bottom of the headlines section).</p>'
            + cbk_html + oil_html + fx_html + african_html
            + foreign_chip + news_html + howto
        )

    def _market_pulse_glossary(self):
        """Plain-English guide for Market Pulse."""
        cards = [
            ("Central Bank Rate (CBR)",
             "The interest rate at which CBK lends to commercial banks. It's the benchmark that flows through to loans, deposits and government bond yields.",
             "CBR rising from 8.75% → 10% usually hurts loan-heavy banks in the short term (borrowers strain) but boosts their bond book. Falling CBR is the reverse."),
            ("Inflation",
             "How fast prices are rising year-over-year. High inflation eats returns and often prompts CBK to raise the CBR.",
             "Inflation 6% while your dividend yield is 5% → your real return is negative unless the share price also rises."),
            ("Oil price (Brent / WTI)",
             "Kenya imports its fuel. Rising oil raises transport, electricity and manufacturing costs — feeds through to inflation and squeezes profit margins.",
             "Brent jumping from $80 → $95 typically weakens the KES and hurts stocks like Bamburi, EABL, KenGen customers."),
            ("KES exchange rates",
             "A weaker KES helps exporters (tea, coffee, tourism) and hurts importers. It also inflates foreign debt costs.",
             "KES weakening from 125 → 135 per USD makes SCOM's tower-lease costs (USD-denominated) more expensive."),
            ("African market comparison",
             "Where the NSE sits vs. its peers (JSE South Africa, NGX Nigeria, EGX Egypt) today. Regional sell-offs often affect NSE via foreign investor flows.",
             "If JSE and EGX are down heavily on a global risk-off day, the NSE often follows a day later."),
            ("Why no 'sentiment' tags?",
             "Reliable sentiment analysis on financial short text requires a paid NLP model. Free keyword-based tagging is wrong ~40% of the time — dangerous when money is involved. So we show the source and let you decide.",
             "A headline like 'Safaricom slides on profit warning' is clearly negative to a human but a keyword tool might miss the context."),
        ]
        items = ''
        for title, what, eg in cards:
            items += (f'<div class="explain-card"><h4>{title}</h4>'
                      f'<p>{what}</p>'
                      f'<p class="eg">📌 <strong>Example:</strong> {eg}</p></div>')
        return ('<div class="section"><h2>📖 How to read this page (plain English)</h2>'
                '<p class="page-intro">What each signal typically means for NSE stocks. '
                'These are heuristics, not rules — every situation has exceptions.</p>'
                f'<div class="explain-grid">{items}</div></div>')

    def _build_dashboard_pages(self, stocks, gainers, losers, sectors, breadth,
                               sector_chart, bullish, bearish, neutral, total,
                               data_date=None, alerts=None, usd_kes=None):
        """
        Build the multi-page dashboard: a clean Overview plus grouped detail
        pages (Technicals, Fundamentals, Dividends, Sectors, Data Quality).
        Every piece of data from the old single page is preserved, just moved
        to a related page. Writes all pages and returns the index.html path.
        """
        now = datetime.now().strftime('%Y-%m-%d %H:%M EAT')
        data_date_str = data_date or datetime.now().strftime('%Y-%m-%d')
        fx = f" · 💵 USD/KES {usd_kes['rate']:.2f}" if (usd_kes and usd_kes.get('rate')) else ""
        subtitle = (f"{now} · {total} stocks · 📅 {data_date_str} · "
                    f"Prices: NSE official close · Fundamentals: TradingView{fx}")

        pv_marker = {
            'ok': ('✓', '#16a34a', 'Verified against independent source'),
            'mismatch': ('❗', '#dc2626', ''),
            'stale': ('🕒', '#d97706', ''),
            'unverified': ('', '#94a3b8', 'No independent source to compare'),
        }

        def price_cell(s):
            price_str = f"{s['price']:.2f}" if s['price'] else '—'
            val = s.get('validation') or {}
            mark, color, tip0 = pv_marker.get(val.get('status', 'unverified'), ('', '#94a3b8', ''))
            tip = val.get('note') or tip0
            mk = f'<span class="pv-mark" style="color:{color}" title="{tip}">{mark}</span>' if mark else ''
            return f'{price_str} {mk}'

        def change_td(s):
            cls = 'positive' if (s['change'] or 0) >= 0 else 'negative'
            txt = f"{s['change']:+.2f}%" if s['change'] is not None else '—'
            return f'<td class="{cls}">{txt}</td>'

        def signal_td(s):
            return f'<td><span class="badge {s["signal_class"]}">{s["signal_label"]}</span></td>'

        def score_td(s):
            sc = s.get('score')
            if sc is None:
                return '<td>—</td>'
            c = 'score-high' if sc >= 70 else 'score-mid' if sc >= 45 else 'score-low'
            return f'<td><span class="score {c}">{sc}</span></td>'

        def sym_td(s):
            link = s['report_file'] if s['report_file'] else '#'
            return f'<td><a href="{link}" class="stock-link"><strong>{s["symbol"]}</strong></a></td>'

        # ---- Market pulse stats (shared on Overview) ----
        breadth_html = ''
        for key, label in [('pct_above_sma50', 'Above SMA50'),
                           ('pct_bullish_macd', 'Bullish MACD'),
                           ('pct_rsi_above_50', 'RSI > 50')]:
            if breadth and key in breadth:
                breadth_html += (f'<div class="stat-card"><div class="stat-value">{breadth[key]}%'
                                 f'</div><div class="stat-label">{label}</div></div>')
        stats_html = (
            '<div class="stats">'
            f'<div class="stat-card"><div class="stat-value">{total}</div><div class="stat-label">Stocks</div></div>'
            f'<div class="stat-card"><div class="stat-value bullish">{bullish}</div><div class="stat-label">Bullish</div></div>'
            f'<div class="stat-card"><div class="stat-value bearish">{bearish}</div><div class="stat-label">Bearish</div></div>'
            f'<div class="stat-card"><div class="stat-value neutral">{neutral}</div><div class="stat-label">Neutral</div></div>'
            f'{breadth_html}</div>'
        )

        # ---- Top movers ----
        gainer_rows = ''.join(
            f'<tr><td>{g["symbol"]}</td><td class="positive">{g["change"]:+.2f}%</td></tr>'
            for g in gainers[:10])
        loser_rows = ''.join(
            f'<tr><td>{l["symbol"]}</td><td class="negative">{l["change"]:+.2f}%</td></tr>'
            for l in losers[:10])
        movers_html = (
            '<div class="grid-2">'
            f'<div class="section"><h2>🟢 Top Gainers</h2><table><tr><th>Symbol</th><th>Change</th></tr>{gainer_rows}</table></div>'
            f'<div class="section"><h2>🔴 Top Losers</h2><table><tr><th>Symbol</th><th>Change</th></tr>{loser_rows}</table></div>'
            '</div>')

        search_bar = ('<div class="filter-bar"><input type="text" id="search" '
                      'placeholder="🔍 Filter by symbol..." oninput="filterTable()"></div>')

        # ---- OVERVIEW page: critical Buy/Sell + price + change + score ----
        ov_rows = ''.join(
            f'<tr>{sym_td(s)}{signal_td(s)}<td>{price_cell(s)}</td>{change_td(s)}{score_td(s)}</tr>'
            for s in stocks)
        overview_body = (
            '<p class="page-intro">Your at-a-glance view: the Buy/Sell signal, price and overall score for every stock. '
            'Use the tabs above for technicals, fundamentals, dividends, sectors and data quality.</p>'
            + stats_html + movers_html +
            '<div class="section"><h2>📋 All Stocks — Signal &amp; Score</h2>'
            + search_bar +
            '<div class="table-wrap"><table id="mainTable"><thead><tr>'
            '<th>Symbol</th><th title="TradingView Buy/Sell rating">TV Signal</th><th>Price</th>'
            '<th>Change</th><th title="0-100 factor screen">Score</th>'
            f'</tr></thead><tbody>{ov_rows}</tbody></table></div></div>')

        # ---- TECHNICALS page ----
        tech_rows = ''
        for s in stocks:
            rsi = f"{s['rsi']:.1f}" if s['rsi'] else '—'
            tech_rows += (
                f'<tr>{sym_td(s)}<td>{price_cell(s)}</td>{change_td(s)}<td>{rsi}</td>'
                f'<td><span class="badge {s["trend"]}">{s["trend"]}</span></td>'
                f'<td><span class="badge {s["ma"]}">{s["ma"].replace("_"," ")}</span></td>'
                f'<td><span class="badge {s["macd"]}">{s["macd"].replace("_"," ")}</span></td>'
                f'<td><span class="badge {s["stochastic"]}">{s["stochastic"]}</span></td>'
                f'<td><span class="badge {s["volume_signal"]}">{s["volume_signal"].replace("_"," ")}</span></td>'
                f'<td><span class="badge {s["overall"]}">{s["overall"]}</span></td></tr>')
        technicals_body = (
            '<p class="page-intro">Momentum &amp; trend indicators for every stock. Green = bullish, red = bearish, amber = neutral.</p>'
            '<div class="section"><h2>📈 Technical Indicators</h2>' + search_bar +
            '<div class="table-wrap"><table id="mainTable"><thead><tr>'
            '<th>Symbol</th><th>Price</th><th>Change</th><th>RSI</th><th>Trend</th>'
            '<th>MA</th><th>MACD</th><th>Stoch</th><th>Vol</th><th>Overall</th>'
            f'</tr></thead><tbody>{tech_rows}</tbody></table></div></div>')

        # ---- FUNDAMENTALS page ----
        def fcls(metric, raw):
            v = self._fund_verdict(metric, raw)
            return {'good': 'fgood', 'mid': 'fmid', 'bad': 'fbad'}.get(v, '')

        fund_rows = ''
        for s in stocks:
            pe = f"{s['pe_ratio']:.1f}" if s['pe_ratio'] else '—'
            peg = f"{s['peg_ratio']:.2f}" if s.get('peg_ratio') else '—'
            pb = f"{s['price_to_book']:.2f}" if s.get('price_to_book') else '—'
            eps = f"{s['eps']:.2f}" if s.get('eps') is not None else '—'
            mcap = self._fmt_mcap(s['market_cap']) if s.get('market_cap') else '—'
            roe = f"{s['roe']:.1f}%" if s.get('roe') is not None else '—'
            nm = f"{s['net_margin']:.1f}%" if s.get('net_margin') is not None else '—'
            de = f"{s['debt_to_equity']:.2f}" if s.get('debt_to_equity') is not None else '—'
            rg = s.get('revenue_growth')
            rg_str = f"{rg:+.1f}%" if rg is not None else '—'
            dy = f"{s['dividend_yield']:.1f}%" if s.get('dividend_yield') else '—'
            fund_rows += (
                f'<tr>{sym_td(s)}<td>{price_cell(s)}</td><td class="mcap-cell">{mcap}</td>'
                f'<td class="{fcls("pe", s.get("pe_ratio"))}">{pe}</td>'
                f'<td class="{fcls("peg", s.get("peg_ratio"))}">{peg}</td>'
                f'<td class="{fcls("pb", s.get("price_to_book"))}">{pb}</td>'
                f'<td class="{fcls("eps", s.get("eps"))}">{eps}</td>'
                f'<td class="{fcls("roe", s.get("roe"))}">{roe}</td>'
                f'<td class="{fcls("nm", s.get("net_margin"))}">{nm}</td>'
                f'<td class="{fcls("de", s.get("debt_to_equity"))}">{de}</td>'
                f'<td class="{fcls("rg", rg)}">{rg_str}</td>'
                f'<td class="{fcls("yield", s.get("dividend_yield"))}">{dy}</td>{score_td(s)}</tr>')
        fundamentals_body = (
            '<p class="page-intro">Valuation, quality, growth &amp; health metrics (from TradingView). '
            '<span class="fgood">Green = good</span>, <span class="fmid">amber = average</span>, '
            '<span class="fbad">red = weak/risky</span>, black = no data. New to these? The plain-English guide with examples is right below the table.</p>'
            '<div class="section"><h2>💰 Fundamentals</h2>' + search_bar +
            '<div class="table-wrap"><table id="mainTable"><thead><tr>'
            '<th>Symbol</th><th>Price</th><th>Market Cap</th><th title="Price / Earnings">P/E</th>'
            '<th title="P/E adjusted for growth">PEG</th><th title="Price / Book value">P/B</th>'
            '<th title="Earnings per share">EPS</th><th title="Return on Equity">ROE</th>'
            '<th title="Net profit margin">Net Margin</th><th title="Debt / Equity">D/E</th>'
            '<th title="Revenue growth vs last year">Rev Growth</th><th>Yield</th><th>Score</th>'
            f'</tr></thead><tbody>{fund_rows}</tbody></table></div></div>'
            + self._fundamentals_explainer())

        # ---- DIVIDENDS page (calendar + table) ----
        today = datetime.now().date()
        cal_rows = []
        for s in stocks:
            ex = s.get('ex_date')
            if not ex:
                continue
            try:
                d = datetime.strptime(ex, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue
            delta = (d - today).days
            if delta < 0:
                cls, when, group, sortk = 'cal-passed', f'{-delta}d ago', 'past', -delta
            elif delta <= 30:
                cls, when, group, sortk = 'cal-near', (f'in {delta}d' if delta else 'today'), 'up', delta
            else:
                cls, when, group, sortk = 'cal-far', f'in {delta}d', 'up', delta
            cal_rows.append({'symbol': s['symbol'], 'dps': s.get('dps'),
                             'yield': s.get('dividend_yield'), 'ex': ex,
                             'cls': cls, 'when': when, 'group': group, 'sortk': sortk})

        def _cal_table(rows, empty_msg):
            if not rows:
                return f'<p class="dq-note">{empty_msg}</p>'
            body = ''
            for r in rows:
                dps = f"{round(r['dps'],2):g}" if r['dps'] else '0'
                yld = f"{r['yield']:.1f}%" if r['yield'] else '—'
                body += (f'<tr><td><strong>{r["symbol"]}</strong></td><td>{dps}</td><td>{yld}</td>'
                         f'<td><span class="cal-chip {r["cls"]}">{r["ex"]}</span></td><td>{r["when"]}</td></tr>')
            return ('<table><thead><tr><th>Symbol</th><th>Div KES</th><th>Yield</th>'
                    f'<th>Ex-Date</th><th>When</th></tr></thead><tbody>{body}</tbody></table>')

        upcoming = sorted([r for r in cal_rows if r['group'] == 'up'], key=lambda r: r['sortk'])
        past = sorted([r for r in cal_rows if r['group'] == 'past'], key=lambda r: r['sortk'])

        def _bc_chip(bc_str):
            # Colour by urgency: red = passed, amber = within a week, green = later.
            if not bc_str:
                return '<span class="exdate-none">—</span>'
            try:
                d = datetime.strptime(bc_str, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                return bc_str
            delta = (d - today).days
            if delta < 0:
                cls = 'bc-passed'
            elif delta <= 7:
                cls = 'bc-soon'
            else:
                cls = 'bc-future'
            return f'<span class="{cls}">{bc_str}</span>'

        def _div_row(s):
            dps = s.get('dps')
            dstatus = s.get('dividend_status')
            if dps and dps > 0:
                if dstatus == 'unverified':
                    amt = f'<span class="div-unverified" title="TradingView figure — not cross-checked">{round(dps,2):g}*</span>'
                else:
                    amt = f'<span class="div-pay">{round(dps,2):g}</span>'
            else:
                amt = '<span class="div-zero">0</span>'
            dy = f"{s['dividend_yield']:.1f}%" if s.get('dividend_yield') else '—'
            bc = _bc_chip(s.get('book_closure') or s.get('ex_date'))
            return f'<tr>{sym_td(s)}<td>{amt}</td><td>{dy}</td><td>{bc}</td></tr>'

        # Dividend payers first (highest yield on top), then the zeros.
        payers = sorted(
            [s for s in stocks if s.get('dps') and s['dps'] > 0],
            key=lambda s: s.get('dividend_yield') or 0, reverse=True)
        nonpayers = [s for s in stocks if not (s.get('dps') and s['dps'] > 0)]
        div_rows = ''.join(_div_row(s) for s in payers + nonpayers)
        dividends_body = (
            '<p class="page-intro">Declared dividends from the NSE calendar (mystocks). '
            '<span class="div-unverified">amber*</span> = TradingView figure that could not be cross-checked.</p>'
            '<div class="section"><h2>💵 Dividend Calendar</h2>'
            '<div class="cal-legend"><span class="cal-chip cal-near">soon (≤30d)</span>'
            '<span class="cal-chip cal-far">later (&gt;30d)</span>'
            '<span class="cal-chip cal-passed">passed</span></div>'
            '<div class="grid-2">'
            f'<div><h3 class="cal-h3">🟢 Upcoming Ex-Dividend Dates <span class="cal-count">({len(upcoming)})</span></h3>'
            f'<div class="table-wrap">{_cal_table(upcoming, "No upcoming ex-dividend dates.")}</div></div>'
            f'<div><h3 class="cal-h3">🔴 Past Ex-Dividend Dates <span class="cal-count">({len(past)})</span></h3>'
            f'<div class="table-wrap">{_cal_table(past, "No past ex-dividend dates recorded.")}</div></div>'
            '</div><div class="dq-note">Own the shares before the book-closure/ex-date to qualify.</div></div>'
            '<div class="section"><h2>📋 All Dividends</h2>'
            '<div class="cal-legend">Book closure / ex-date: '
            '<span class="bc-future">upcoming</span>'
            '<span class="bc-soon">within a week</span>'
            '<span class="bc-passed">passed</span></div>'
            + search_bar +
            '<div class="table-wrap"><table id="mainTable"><thead><tr>'
            '<th>Symbol</th><th>Div KES</th><th>Yield</th><th>Book closure / Ex-date</th>'
            f'</tr></thead><tbody>{div_rows}</tbody></table></div></div>')

        # ---- SECTORS page ----
        sector_cards = ''
        if sectors:
            for name, data in sectors.items():
                chg = data['avg_change_pct']
                cls = 'positive' if chg >= 0 else 'negative'
                sector_cards += (f'<div class="sector-card"><h3>{name}</h3>'
                                 f'<div class="sector-change {cls}">{chg:+.2f}%</div>'
                                 f'<div class="sector-detail">{data["count"]} stocks | RSI '
                                 f'{data.get("avg_rsi", "—")} | {data["bullish_ratio"]}% bullish</div></div>')
        sector_chart_html = (f'<img src="data:image/png;base64,{sector_chart}" class="chart-img" '
                             'alt="Sector Performance">') if sector_chart else ''
        sectors_body = (
            '<p class="page-intro">How each NSE sector performed today.</p>'
            f'<div class="section"><h2>📊 Sector Performance</h2><div class="grid-3">{sector_cards}</div>'
            f'{sector_chart_html}</div>')

        # ---- DATA QUALITY page (validation + alerts) ----
        v_ok = v_mismatch = v_stale = v_unverified = 0
        mismatch_list = []
        for s in stocks:
            st = (s.get('validation') or {}).get('status')
            if st == 'ok':
                v_ok += 1
            elif st == 'mismatch':
                v_mismatch += 1
                mismatch_list.append(s)
            elif st == 'stale':
                v_stale += 1
            else:
                v_unverified += 1
        mismatch_note = ''
        if mismatch_list:
            items = ', '.join(
                f"{m['symbol']} ({(m.get('validation') or {}).get('pct_diff'):+.1f}%)"
                for m in mismatch_list)
            mismatch_note = (f'<div class="dq-note dq-mismatch">⚠️ TradingView differs from the NSE '
                             f'official close for: {items}</div>')
        alerts_cards = ''
        for sym in sorted((alerts or {}).keys()):
            items = alerts[sym]
            if items:
                alerts_cards += (f'<div class="alert-card"><div class="sym">{sym}</div>'
                                 f'<div class="items">{"<br>".join(items)}</div></div>')
        alerts_section = (f'<div class="section"><h2>🔔 Alerts &amp; Signals</h2>'
                          f'<div class="alerts-grid">{alerts_cards}</div></div>') if alerts_cards else ''
        quality_body = (
            '<p class="page-intro">How much to trust today\'s prices, and notable per-stock alerts.</p>'
            '<div class="section"><h2>✅ Data Quality</h2><div class="stats">'
            f'<div class="stat-card"><div class="stat-value bullish">{v_ok}</div><div class="stat-label">Verified</div></div>'
            f'<div class="stat-card"><div class="stat-value bearish">{v_mismatch}</div><div class="stat-label">Price mismatch</div></div>'
            f'<div class="stat-card"><div class="stat-value neutral">{v_stale}</div><div class="stat-label">Stale / thin</div></div>'
            f'<div class="stat-card"><div class="stat-value">{v_unverified}</div><div class="stat-label">Unverified</div></div></div>'
            '<div class="dq-note">Prices shown are the <strong>NSE official close</strong> (afx.kwayisi.org), '
            'cross-checked against TradingView. ✓ = TradingView confirms it · ❗ = differs · 🕒 = last traded &gt;1 day ago.</div>'
            f'{mismatch_note}</div>{alerts_section}')

        # ---- NEXT EARNINGS page ----
        # Only stocks whose next earnings date is TODAY OR IN THE FUTURE
        # (in Nairobi time). Nearest date first. Skip stocks with no earnings
        # date entirely — this page is about upcoming events to watch.
        try:
            from zoneinfo import ZoneInfo
            today_ke = datetime.now(ZoneInfo("Africa/Nairobi")).date()
        except Exception:
            today_ke = datetime.now().date()

        earnings_rows_raw = []
        for s in stocks:
            nd = s.get('earnings_next_date')
            if not nd:
                continue
            try:
                edate = datetime.strptime(nd, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue
            if edate < today_ke:
                continue  # past — exclude
            days = (edate - today_ke).days
            earnings_rows_raw.append((days, edate, s))
        earnings_rows_raw.sort(key=lambda t: t[0])

        # Urgency chip: red = today, amber = ≤7 days, green = further out.
        def _earnings_chip(days, date_str):
            if days == 0:
                cls, when = 'bc-passed', 'today'
            elif days <= 7:
                cls, when = 'bc-soon', f'in {days}d'
            else:
                cls, when = 'bc-future', f'in {days}d'
            return f'<span class="{cls}">{date_str}</span>', when

        earnings_rows = ''
        for days, edate, s in earnings_rows_raw:
            chip, when = _earnings_chip(days, edate.strftime('%Y-%m-%d'))
            price_str = f"{s['price']:.2f}" if s['price'] else '—'
            earnings_rows += (
                f'<tr>{sym_td(s)}<td>{chip}</td><td>{when}</td>'
                f'<td>{price_str}</td>{change_td(s)}{signal_td(s)}{score_td(s)}</tr>')
        if not earnings_rows:
            earnings_rows = ('<tr><td colspan="7" style="color:#94a3b8; text-align:center;">'
                             'No upcoming earnings dates on record.</td></tr>')
        earnings_body = (
            '<p class="page-intro">Every stock with an upcoming earnings-release date, '
            '<strong>nearest first</strong>. Stocks with no scheduled date, or a date already '
            'in the past, are excluded. Earnings dates come from TradingView — confirm on the '
            'NSE announcement before making decisions.</p>'
            '<div class="section"><h2>📅 Upcoming Earnings</h2>'
            '<div class="cal-legend">'
            '<span class="bc-passed">today</span>'
            '<span class="bc-soon">within a week</span>'
            '<span class="bc-future">later</span>'
            '</div>'
            + search_bar +
            '<div class="table-wrap"><table id="mainTable"><thead><tr>'
            '<th>Symbol</th><th>Earnings Date</th><th>When</th>'
            '<th>Price</th><th>Change</th><th>TV Signal</th><th>Score</th>'
            f'</tr></thead><tbody>{earnings_rows}</tbody></table></div>'
            f'<div class="dq-note">{len(earnings_rows_raw)} stock(s) with an upcoming '
            'earnings release. Earnings dates are only published for a subset of NSE stocks; '
            'a stock not listed here simply has no scheduled date in the feed.</div></div>')

        # ---- FOREIGN FLOWS page (manual weekly input) ----
        foreign_body = self._build_foreign_flows_body(sym_td)

        # ---- MARKET PULSE page (news + CBK + oil + FX + African indices) ----
        pulse_body = self._build_market_pulse_body(
            analysis_results_for_pulse=(
                {s['symbol']: {'daily_change_pct': s.get('change')} for s in stocks}
            )
        )

        # ---- Assemble & write all pages ----
        pages = {
            'index.html': self._page_shell('NSE Dashboard — Overview', 'index.html', subtitle, overview_body, with_filter=True),
            'technicals.html': self._page_shell('NSE — Technicals', 'technicals.html', subtitle, technicals_body, with_filter=True),
            'fundamentals.html': self._page_shell('NSE — Fundamentals', 'fundamentals.html', subtitle, fundamentals_body, with_filter=True),
            'dividends.html': self._page_shell('NSE — Dividends', 'dividends.html', subtitle, dividends_body, with_filter=True),
            'earnings.html': self._page_shell('NSE — Next Earnings', 'earnings.html', subtitle, earnings_body, with_filter=True),
            'sectors.html': self._page_shell('NSE — Sectors', 'sectors.html', subtitle, sectors_body),
            'foreign.html': self._page_shell('NSE — Foreign Flows', 'foreign.html', subtitle, foreign_body),
            'pulse.html': self._page_shell('NSE — Market Pulse', 'pulse.html', subtitle, pulse_body),
            'quality.html': self._page_shell('NSE — Data Quality', 'quality.html', subtitle, quality_body),
        }
        for filename, html in pages.items():
            with open(os.path.join(self.output_dir, filename), 'w', encoding='utf-8') as f:
                f.write(html)
        logger.info(f"Dashboard saved: {len(pages)} pages — {', '.join(pages.keys())}")
        return os.path.join(self.output_dir, 'index.html')

    def _build_index_html(self, stocks, gainers, losers, sectors, breadth,
                          sector_chart, bullish, bearish, neutral, total,
                          data_date=None, alerts=None, usd_kes=None):
        """[DEPRECATED — replaced by _build_dashboard_pages] Kept for reference."""
        now = datetime.now().strftime('%Y-%m-%d %H:%M EAT')
        data_date_str = data_date or datetime.now().strftime('%Y-%m-%d')

        # ---- Data-quality summary (price validation) ----
        v_ok = v_mismatch = v_stale = v_unverified = 0
        mismatch_list = []
        for s in stocks:
            st = (s.get('validation') or {}).get('status')
            if st == 'ok':
                v_ok += 1
            elif st == 'mismatch':
                v_mismatch += 1
                mismatch_list.append(s)
            elif st == 'stale':
                v_stale += 1
            else:
                v_unverified += 1

        # Price-validation marker per status
        pv_marker = {
            'ok': ('✓', '#16a34a', 'Verified against independent source'),
            'mismatch': ('❗', '#dc2626', ''),
            'stale': ('🕒', '#d97706', ''),
            'unverified': ('', '#94a3b8', 'No independent source to compare'),
        }

        # Build stock rows
        stock_rows = ''
        for s in stocks:
            chg_class = 'positive' if (s['change'] or 0) >= 0 else 'negative'
            chg_str = f"{s['change']:+.2f}%" if s['change'] is not None else '—'
            price_str = f"{s['price']:.2f}" if s['price'] else '—'
            rsi_str = f"{s['rsi']:.1f}" if s['rsi'] else '—'
            pe_str = f"{s['pe_ratio']:.1f}" if s['pe_ratio'] else '—'
            mcap_str = self._fmt_mcap(s['market_cap']) if s.get('market_cap') else '—'
            link = s['report_file'] if s['report_file'] else '#'

            # Dividend yield
            dy = s.get('dividend_yield')
            dy_str = f"{dy:.1f}%" if dy else '—'

            # Dividend amount (KES/share) — 0 when the stock pays nothing
            dps = s.get('dps')
            if dps and dps > 0:
                div_html = f'<span class="div-pay">{dps:g}</span>'
            else:
                div_html = '<span class="div-zero">0</span>'

            # Ex-dividend date — upcoming (still buyable) vs past vs none
            ex_date = s.get('ex_date')
            if ex_date and s.get('ex_upcoming'):
                exdate_html = f'<span class="exdate-upcoming" title="Buy before this date to receive the dividend">{ex_date}</span>'
            elif ex_date:
                exdate_html = f'<span class="exdate-past" title="Most recent ex-dividend date (already passed)">{ex_date}</span>'
            else:
                exdate_html = '<span class="exdate-none">—</span>'

            # Score badge (colour by band)
            score = s.get('score')
            if score is None:
                score_html = '—'
            else:
                sc_class = 'score-high' if score >= 70 else 'score-mid' if score >= 45 else 'score-low'
                score_html = f'<span class="score {sc_class}">{score}</span>'

            # Price-validation marker (with tooltip)
            val = s.get('validation') or {}
            status = val.get('status', 'unverified')
            mark, color, default_tip = pv_marker.get(status, ('', '#94a3b8', ''))
            tip = val.get('note') or default_tip
            mark_html = (f'<span class="pv-mark" style="color:{color}" title="{tip}">{mark}</span>'
                         if mark else '')

            stock_rows += f'''
            <tr>
                <td><a href="{link}" class="stock-link"><strong>{s['symbol']}</strong></a></td>
                <td><span class="badge {s['signal_class']}">{s['signal_label']}</span></td>
                <td>{price_str} {mark_html}</td>
                <td class="{chg_class}">{chg_str}</td>
                <td>{dy_str}</td>
                <td>{div_html}</td>
                <td>{exdate_html}</td>
                <td>{pe_str}</td>
                <td class="mcap-cell">{mcap_str}</td>
                <td>{rsi_str}</td>
                <td><span class="badge {s['trend']}">{s['trend']}</span></td>
                <td><span class="badge {s['ma']}">{s['ma'].replace('_',' ')}</span></td>
                <td><span class="badge {s['macd']}">{s['macd'].replace('_',' ')}</span></td>
                <td><span class="badge {s['stochastic']}">{s['stochastic']}</span></td>
                <td><span class="badge {s['volume_signal']}">{s['volume_signal'].replace('_',' ')}</span></td>
                <td><span class="badge {s['overall']}">{s['overall']}</span></td>
                <td>{score_html}</td>
            </tr>'''

        # Build gainer/loser rows
        gainer_rows = ''.join(
            f'<tr><td>{g["symbol"]}</td><td class="positive">{g["change"]:+.2f}%</td></tr>'
            for g in gainers[:10]
        )
        loser_rows = ''.join(
            f'<tr><td>{l["symbol"]}</td><td class="negative">{l["change"]:+.2f}%</td></tr>'
            for l in losers[:10]
        )

        # Build sector cards
        sector_cards = ''
        if sectors:
            for name, data in sectors.items():
                chg = data['avg_change_pct']
                cls = 'positive' if chg >= 0 else 'negative'
                sector_cards += f'''
                <div class="sector-card">
                    <h3>{name}</h3>
                    <div class="sector-change {cls}">{chg:+.2f}%</div>
                    <div class="sector-detail">{data['count']} stocks | RSI {data.get('avg_rsi', '—')} | {data['bullish_ratio']}% bullish</div>
                </div>'''

        # Breadth stats
        breadth_html = ''
        if breadth:
            for key, label in [
                ('pct_above_sma50', 'Above SMA50'),
                ('pct_bullish_macd', 'Bullish MACD'),
                ('pct_rsi_above_50', 'RSI > 50'),
            ]:
                if key in breadth:
                    breadth_html += f'''
                    <div class="stat-card">
                        <div class="stat-value">{breadth[key]}%</div>
                        <div class="stat-label">{label}</div>
                    </div>'''

        sector_chart_html = ''
        if sector_chart:
            sector_chart_html = f'<img src="data:image/png;base64,{sector_chart}" class="chart-img" alt="Sector Performance">'

        # ---- Alerts section ----
        alerts_html = ''
        if alerts:
            cards = ''
            for sym in sorted(alerts.keys()):
                items = alerts[sym]
                if not items:
                    continue
                items_html = '<br>'.join(items)
                cards += (f'<div class="alert-card"><div class="sym">{sym}</div>'
                          f'<div class="items">{items_html}</div></div>')
            if cards:
                alerts_html = f'''
<div class="section">
    <h2>🔔 Alerts &amp; Signals</h2>
    <div class="alerts-grid">{cards}</div>
</div>'''

        # ---- Data-quality section (price validation) ----
        dq_html = ''
        if (v_ok + v_mismatch + v_stale + v_unverified) > 0:
            mismatch_note = ''
            if mismatch_list:
                items = ', '.join(
                    f"{m['symbol']} ({(m.get('validation') or {}).get('pct_diff'):+.1f}%)"
                    for m in mismatch_list
                )
                mismatch_note = (f'<div class="dq-note dq-mismatch">⚠️ TradingView differs from '
                                 f'the NSE official close for: {items}</div>')
            dq_html = f'''
<div class="section">
    <h2>✅ Data Quality</h2>
    <div class="stats">
        <div class="stat-card"><div class="stat-value bullish">{v_ok}</div><div class="stat-label">Verified</div></div>
        <div class="stat-card"><div class="stat-value bearish">{v_mismatch}</div><div class="stat-label">Price mismatch</div></div>
        <div class="stat-card"><div class="stat-value neutral">{v_stale}</div><div class="stat-label">Stale / thin</div></div>
        <div class="stat-card"><div class="stat-value">{v_unverified}</div><div class="stat-label">Unverified</div></div>
    </div>
    <div class="dq-note">Prices shown are the <strong>NSE official close</strong> (afx.kwayisi.org), cross-checked against TradingView.
    ✓ = TradingView confirms it · ❗ = TradingView differs (price uncertain) · 🕒 = last traded &gt;1 day ago.</div>
    {mismatch_note}
</div>'''

        # ---- FX chip for the header ----
        fx_html = ''
        if usd_kes and usd_kes.get('rate'):
            fx_html = f" · 💵 USD/KES {usd_kes['rate']:.2f}"

        # ---- Dividend calendar (ex-dividend dates, colour-coded by proximity) ----
        today = datetime.now().date()
        cal_rows = []
        for s in stocks:
            ex = s.get('ex_date')
            if not ex:
                continue
            try:
                d = datetime.strptime(ex, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                continue
            delta = (d - today).days
            if delta < 0:
                cls, when, group, sortk = 'cal-passed', f'{-delta}d ago', 'past', -delta
            elif delta <= 30:
                cls, when, group, sortk = 'cal-near', (f'in {delta}d' if delta else 'today'), 'up', delta
            else:
                cls, when, group, sortk = 'cal-far', f'in {delta}d', 'up', delta
            cal_rows.append({
                'symbol': s['symbol'], 'dps': s.get('dps'),
                'yield': s.get('dividend_yield'), 'ex': ex,
                'cls': cls, 'when': when, 'group': group, 'sortk': sortk,
            })

        def _cal_table(rows, empty_msg):
            if not rows:
                return f'<p class="dq-note">{empty_msg}</p>'
            body = ''
            for r in rows:
                dps = f"{r['dps']:g}" if r['dps'] else '0'
                yld = f"{r['yield']:.1f}%" if r['yield'] else '—'
                body += (f'<tr><td><strong>{r["symbol"]}</strong></td>'
                         f'<td>{dps}</td><td>{yld}</td>'
                         f'<td><span class="cal-chip {r["cls"]}">{r["ex"]}</span></td>'
                         f'<td>{r["when"]}</td></tr>')
            return ('<table><thead><tr><th>Symbol</th><th>Div KES</th><th>Yield</th>'
                    f'<th>Ex-Date</th><th>When</th></tr></thead><tbody>{body}</tbody></table>')

        upcoming = sorted([r for r in cal_rows if r['group'] == 'up'], key=lambda r: r['sortk'])
        past = sorted([r for r in cal_rows if r['group'] == 'past'], key=lambda r: r['sortk'])
        dividend_calendar_html = f'''
<div class="section">
    <h2>💵 Dividend Calendar</h2>
    <div class="cal-legend">
        <span class="cal-chip cal-near">soon (≤30d)</span>
        <span class="cal-chip cal-far">later (&gt;30d)</span>
        <span class="cal-chip cal-passed">passed</span>
    </div>
    <div class="grid-2">
        <div>
            <h3 class="cal-h3">🟢 Upcoming Ex-Dividend Dates <span class="cal-count">({len(upcoming)})</span></h3>
            <div class="table-wrap">{_cal_table(upcoming, "No upcoming ex-dividend dates in the current data.")}</div>
        </div>
        <div>
            <h3 class="cal-h3">🔴 Past Ex-Dividend Dates <span class="cal-count">({len(past)})</span></h3>
            <div class="table-wrap">{_cal_table(past, "No past ex-dividend dates recorded.")}</div>
        </div>
    </div>
    <div class="dq-note">Buy <strong>before</strong> a green/yellow ex-date to receive that dividend. Div KES = dividend per share for the year (0 = none).</div>
    <div class="dq-note dq-mismatch">⚠️ Dividend <strong>payment dates</strong> are not published by our data feed (TradingView provides ex-dividend dates only, and free NSE sources checked were stale). On the NSE, payment typically follows the ex-date by ~3–8 weeks — confirm the exact date in the company's official NSE announcement.</div>
</div>'''

        return f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>NSE Daily Dashboard — {now}</title>
<style>
* {{ margin: 0; padding: 0; box-sizing: border-box; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #f1f5f9; color: #1e293b; }}
.container {{ max-width: 1400px; margin: 0 auto; padding: 20px; }}

/* Header */
.header {{ background: linear-gradient(135deg, #0f172a, #1e293b); color: white; padding: 30px; border-radius: 12px; margin-bottom: 20px; text-align: center; }}
.header h1 {{ font-size: 2rem; margin-bottom: 5px; }}
.header .date {{ color: #94a3b8; font-size: 0.9rem; }}

/* Stats row */
.stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(130px, 1fr)); gap: 12px; margin-bottom: 20px; }}
.stat-card {{ background: white; padding: 16px; border-radius: 10px; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
.stat-card .stat-value {{ font-size: 1.8rem; font-weight: 700; }}
.stat-card .stat-label {{ font-size: 0.75rem; color: #64748b; text-transform: uppercase; margin-top: 4px; }}
.stat-card .bullish {{ color: #22c55e; }}
.stat-card .bearish {{ color: #ef4444; }}
.stat-card .neutral {{ color: #f59e0b; }}

/* Section */
.section {{ background: white; border-radius: 10px; padding: 20px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }}
.section h2 {{ font-size: 1.1rem; margin-bottom: 16px; padding-bottom: 8px; border-bottom: 2px solid #3b82f6; display: inline-block; }}

/* Table */
.table-wrap {{ overflow-x: auto; }}
table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
th, td {{ padding: 8px 10px; text-align: left; border-bottom: 1px solid #e2e8f0; white-space: nowrap; }}
th {{ background: #f8fafc; color: #64748b; font-size: 0.7rem; text-transform: uppercase; font-weight: 600; position: sticky; top: 0; }}
tr:hover {{ background: #f8fafc; }}
.stock-link {{ color: #3b82f6; text-decoration: none; font-weight: 600; }}
.stock-link:hover {{ text-decoration: underline; }}

/* Badges */
.badge {{ padding: 2px 8px; border-radius: 10px; font-size: 0.7rem; font-weight: 600; text-transform: capitalize; }}
.bullish, .golden_cross, .bullish_cross, .oversold, .buy {{ background: #dcfce7; color: #166534; }}
.bearish, .death_cross, .bearish_cross, .overbought, .sell {{ background: #fee2e2; color: #991b1b; }}
.neutral, .within_bands, .normal {{ background: #fef3c7; color: #92400e; }}
.strong_buy {{ background: #16a34a; color: #ffffff; }}
.strong_sell {{ background: #dc2626; color: #ffffff; }}

/* Score chips */
.score {{ display: inline-block; min-width: 30px; padding: 2px 8px; border-radius: 10px; font-weight: 700; font-size: 0.75rem; text-align: center; }}
.score-high {{ background: #dcfce7; color: #166534; }}
.score-mid {{ background: #fef3c7; color: #92400e; }}
.score-low {{ background: #fee2e2; color: #991b1b; }}
.pv-mark {{ font-size: 0.75rem; cursor: help; }}

/* Dividend amount (teal "money" highlight) vs 0 (muted) */
.div-pay {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #ccfbf1; color: #0f766e; font-weight: 700; }}
.div-zero {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #f1f5f9; color: #94a3b8; font-weight: 600; }}
/* Ex-dividend date — a different colour family from the amount */
.exdate-upcoming {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #16a34a; color: #ffffff; font-weight: 700; cursor: help; }}
.exdate-past {{ display: inline-block; padding: 2px 8px; border-radius: 10px; background: #e0e7ff; color: #3730a3; font-weight: 600; cursor: help; }}
.exdate-none {{ color: #cbd5e1; }}

/* Dividend calendar chips: green=soon, yellow=later, red=passed */
.cal-chip {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-weight: 700; font-size: 0.78rem; }}
.cal-near {{ background: #16a34a; color: #ffffff; }}
.cal-far {{ background: #fde68a; color: #92400e; }}
.cal-passed {{ background: #fecaca; color: #991b1b; }}
.cal-legend {{ display: flex; gap: 8px; flex-wrap: wrap; margin-bottom: 14px; }}
.cal-h3 {{ font-size: 0.95rem; margin-bottom: 10px; }}
.cal-count {{ color: #94a3b8; font-weight: 400; }}

/* Alerts */
.alerts-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 10px; }}
.alert-card {{ background: #f8fafc; border: 1px solid #e2e8f0; border-left: 3px solid #3b82f6; border-radius: 8px; padding: 10px 12px; }}
.alert-card .sym {{ font-weight: 700; color: #3b82f6; margin-bottom: 4px; }}
.alert-card .items {{ font-size: 0.8rem; color: #475569; line-height: 1.5; }}
.dq-note {{ font-size: 0.8rem; color: #64748b; margin-top: 8px; }}
.dq-mismatch {{ color: #991b1b; }}
.undefined {{ background: #f1f5f9; color: #64748b; }}
.high_volume {{ background: #ede9fe; color: #5b21b6; }}
.low_volume {{ background: #f1f5f9; color: #64748b; }}

/* Market cap */
.mcap-cell {{ font-size: 0.8rem; color: #475569; white-space: nowrap; }}

/* Changes */
.positive {{ color: #22c55e; font-weight: 600; }}
.negative {{ color: #ef4444; font-weight: 600; }}

/* Grid layouts */
.grid-2 {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
.grid-3 {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(200px, 1fr)); gap: 12px; }}

/* Sector cards */
.sector-card {{ background: #f8fafc; padding: 14px; border-radius: 8px; border: 1px solid #e2e8f0; }}
.sector-card h3 {{ font-size: 0.9rem; margin-bottom: 4px; }}
.sector-change {{ font-size: 1.3rem; font-weight: 700; }}
.sector-detail {{ font-size: 0.75rem; color: #64748b; margin-top: 4px; }}

/* Chart */
.chart-img {{ max-width: 100%; border-radius: 8px; margin-top: 12px; }}

/* Filter */
.filter-bar {{ margin-bottom: 16px; display: flex; gap: 8px; flex-wrap: wrap; }}
.filter-bar input {{ padding: 8px 12px; border: 1px solid #e2e8f0; border-radius: 6px; font-size: 0.9rem; width: 200px; }}
.filter-bar select {{ padding: 8px 12px; border: 1px solid #e2e8f0; border-radius: 6px; font-size: 0.9rem; }}

/* Footer */
.footer {{ text-align: center; padding: 20px; color: #94a3b8; font-size: 0.8rem; }}

@media (max-width: 768px) {{ .grid-2 {{ grid-template-columns: 1fr; }} }}
</style>
</head>
<body>
<div class="container">

<div class="header">
    <h1>🇰🇪 NSE Daily Dashboard</h1>
    <div class="date">{now} · {total} stocks analyzed · 📅 Financial data: {data_date_str} · Prices: NSE official close · Fundamentals: TradingView{fx_html}</div>
</div>

<!-- Market Stats -->
<div class="stats">
    <div class="stat-card"><div class="stat-value">{total}</div><div class="stat-label">Stocks</div></div>
    <div class="stat-card"><div class="stat-value bullish">{bullish}</div><div class="stat-label">Bullish</div></div>
    <div class="stat-card"><div class="stat-value bearish">{bearish}</div><div class="stat-label">Bearish</div></div>
    <div class="stat-card"><div class="stat-value neutral">{neutral}</div><div class="stat-label">Neutral</div></div>
    {breadth_html}
</div>

{dq_html}

{dividend_calendar_html}

{alerts_html}

<!-- Sector Performance -->
<div class="section">
    <h2>📊 Sector Performance</h2>
    <div class="grid-3">{sector_cards}</div>
    {sector_chart_html}
</div>

<!-- Top Movers -->
<div class="grid-2">
    <div class="section">
        <h2>🟢 Top Gainers</h2>
        <table><tr><th>Symbol</th><th>Change</th></tr>{gainer_rows}</table>
    </div>
    <div class="section">
        <h2>🔴 Top Losers</h2>
        <table><tr><th>Symbol</th><th>Change</th></tr>{loser_rows}</table>
    </div>
</div>

<!-- All Stocks -->
<div class="section">
    <h2>📋 All Stocks</h2>
    <div class="filter-bar">
        <input type="text" id="search" placeholder="🔍 Filter stocks..." oninput="filterTable()">
        <select id="signalFilter" onchange="filterTable()">
            <option value="">All Signals</option>
            <option value="bullish">Bullish</option>
            <option value="bearish">Bearish</option>
            <option value="neutral">Neutral</option>
        </select>
    </div>
    <div class="table-wrap">
        <table id="stockTable">
            <thead>
                <tr>
                    <th>Symbol</th>
                    <th title="TradingView technical rating — Buy / Sell / Neutral">TV Signal</th>
                    <th>Price</th><th>Change</th>
                    <th title="Dividend yield">Yield</th>
                    <th title="Dividend per share (KES). 0 = no dividend">Div KES</th>
                    <th title="Ex-dividend date. Green = upcoming (buy before it to receive the dividend)">Ex-Div Date</th>
                    <th>P/E</th><th>Market Cap</th>
                    <th>RSI</th><th>Trend</th><th>MA</th><th>MACD</th><th>Stoch</th><th>Vol</th>
                    <th>Overall</th>
                    <th title="Transparent 0-100 factor screen (value, quality, momentum, dividend, liquidity)">Score</th>
                </tr>
            </thead>
            <tbody>{stock_rows}</tbody>
        </table>
    </div>
</div>

<div class="footer">Generated by Kenyan Stock Analyzer · Click any stock symbol to view detailed report</div>

</div>

<script>
function filterTable() {{
    const search = document.getElementById('search').value.toLowerCase();
    const signal = document.getElementById('signalFilter').value.toLowerCase();
    const rows = document.querySelectorAll('#stockTable tbody tr');
    rows.forEach(row => {{
        const text = row.textContent.toLowerCase();
        const overall = row.querySelector('.badge:last-child')?.textContent.toLowerCase() || '';
        const matchSearch = !search || text.includes(search);
        const matchSignal = !signal || overall.includes(signal);
        row.style.display = (matchSearch && matchSignal) ? '' : 'none';
    }});
}}
</script>
</body>
</html>'''
        """Render a Jinja2 template, falling back to inline if file missing."""
        try:
            template = self.env.get_template(template_name)
            return template.render(**data)
        except Exception as e:
            logger.warning(f"Template {template_name} not found: {e}")
            return self._fallback_html(template_name, data)

    def _render(self, template_name, data):
        """Render a Jinja2 template, falling back to inline if file missing."""
        try:
            template = self.env.get_template(template_name)
            return template.render(**data)
        except Exception as e:
            logger.warning(f"Template {template_name} not found: {e}")
            return self._fallback_html(template_name, data)

    def _fallback_html(self, template_name, data):
        """Minimal fallback HTML when template files are missing."""
        return f"""<!DOCTYPE html><html><head><title>{template_name}</title>
        <style>body{{font-family:sans-serif;margin:40px;}}
        table{{border-collapse:collapse;width:100%}}
        td,th{{border:1px solid #ddd;padding:8px;text-align:left}}
        .bullish{{color:green}}.bearish{{color:red}}</style></head>
        <body><h1>{template_name}</h1><p>Generated at {data.get('generated_at', 'N/A')}</p>
        <pre>{json.dumps({k: str(v) for k, v in data.items() if k != 'charts'}, indent=2, default=str)}</pre>
        </body></html>"""

    def _save_report(self, prefix, html_content, report_type):
        """
        Save HTML content and optionally generate PDF.

        Returns:
            str or tuple of path(s).
        """
        html_path = os.path.join(
            self.output_dir, f"{prefix}_{self.timestamp}.html"
        )
        with open(html_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        logger.info(f"Saved HTML: {html_path}")

        if report_type == 'html':
            return html_path

        if report_type in ('pdf', 'both'):
            pdf_path = self._generate_pdf(html_content, prefix)
            if pdf_path:
                return (html_path, pdf_path) if report_type == 'both' else pdf_path
            return html_path if report_type == 'both' else None

        return html_path

    def _generate_pdf(self, html_content, prefix):
        """Generate PDF from HTML, gracefully handling failures."""
        if not WEASYPRINT_AVAILABLE:
            logger.warning(
                "PDF generation skipped — WeasyPrint not available. "
                "Run: brew install pango glib"
            )
            return None

        try:
            pdf_path = os.path.join(
                self.output_dir, f"{prefix}_{self.timestamp}.pdf"
            )
            HTML(string=html_content).write_pdf(pdf_path)
            logger.info(f"Saved PDF: {pdf_path}")
            return pdf_path
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            return None


# ---- Test ----
if __name__ == "__main__":
    from logger import setup_logging
    setup_logging()
    logger = get_logger(__name__)

    # Generate sample data
    dates = pd.date_range('2025-06-01', periods=60, freq='B')
    np.random.seed(42)
    close = np.random.randn(60).cumsum() + 100
    data = pd.DataFrame({
        'open': close + np.random.randn(60) * 0.5,
        'high': close + abs(np.random.randn(60)) * 2,
        'low': close - abs(np.random.randn(60)) * 2,
        'close': close,
        'volume': np.random.randint(50000, 500000, 60),
    }, index=dates)

    # Quick analysis
    from analysis_engine import AnalysisEngine
    engine = AnalysisEngine()
    result = engine.analyze_stock(data)

    # Generate report
    rg = ReportGenerator()
    path = rg.generate_stock_report('SCOM', result, report_type='html')
    print(f"\nReport: {path}")
    print("Open it in your browser to verify!")